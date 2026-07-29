# -*- coding: utf-8 -*-
"""
J-5 산출물 일괄 감사 — 배포 전 반드시 통과해야 하는 게이트

배경
  워크북을 만들고 실제로 열어보지 않아 오류를 세 번 놓쳤다.
    ① 계산 시트가 빈칸으로 보임 (openpyxl 이 캐시값을 안 남김)
    ② PERCENTILE.INC 가 #NAME? (Excel 2010+ 함수는 _xlfn. 접두어 필요)
    ③ 대조 항목 12개 중 A(자료마감일) 누락, K·L 단위 불일치
  전부 파일을 한 번 열어 보면 즉시 드러나는 것들이었다.
  그래서 눈으로 확인하는 대신 이 스크립트를 통과시킨다.

검사 5종
  1) 블라인드    수기 워크시트에 코드 산출값이 남아 있지 않은가
  2) 항목 정합   워크시트·봉인·엑셀본의 대조 항목이 같은가
  3) 가이드 주소 기입가이드가 인용한 셀이 실제 입력칸인가
  4) 값 일치     독립 재계산 · 봉인 · 엑셀 세 값이 같은가
  5) 수식 오류   엑셀본에 #NAME? 등이 없는가

실행
  python audit_j5.py
  종료코드 0 = 통과 / 1 = 실패
"""
import os
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
import paths as P  # noqa: E402

P.force_utf8_stdout()
PUB = os.path.join(P.DATA, "pilot_run", "output_f1")
INP = os.path.join(P.DATA, "pilot_run", "input_krxbm")
WS = os.path.join(HERE, "J5_수기검산_워크시트.xlsx")
SEALED = os.path.join(HERE, "J5_코드값_봉인.xlsx")
XLS = os.path.join(HERE, "J5_엑셀수식_교차검산.xlsx")
GUIDE = os.path.join(HERE, "기입가이드_J5.md")
ERRS = ("#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")

fails = []


def check(name, ok, detail=""):
    print(f"  [{'통과' if ok else '실패'}] {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        fails.append(name)


# ── 독립 재계산: 산출물이 아니라 원자료에서 다시 만든다 ──────────
def truth():
    st = pd.read_csv(os.path.join(PUB, "daily_market_state.csv"),
                     dtype={"security_id": str}, parse_dates=["market_date"])
    cal = pd.read_csv(os.path.join(INP, "calendar.csv"), parse_dates=["market_date"])
    fx = pd.read_csv(os.path.join(INP, "fx.csv"), parse_dates=["market_date"]).set_index("market_date")["fx_rate"]
    cons = pd.read_csv(os.path.join(PUB, "constituents_2026-03-31.csv"), dtype={"security_id": str})

    kr = set(cal[(cal.market == "KR") & (cal.is_market_open == 1)].market_date)
    us = set(cal[(cal.market == "US") & (cal.is_market_open == 1)].market_date)
    cut = [d for d in sorted(kr & us) if d <= pd.Timestamp("2026-06-30")][-6]
    win = [d for d in sorted(us) if d <= pd.Timestamp("2026-06-23")][-90:]

    adtv = {}
    for s in ["ALAB", "ANET", "APH", "ATI", "ETN", "GEV", "KTOS", "TER"]:
        sub = st[(st.security_id == s) & (st.market_date.isin(win))]
        adtv[s] = float((sub.raw_close.fillna(0) * sub.volume.fillna(0)).sum()) / 90
    v = np.sort(list(adtv.values()))
    p10 = float(np.percentile(v, 10))

    sc = cons[cons.selected_flag == 1]
    cnt = sc.groupby("cell_id").size().to_dict()
    wt = {r.security_id: (1 / 6) / cnt[r.cell_id] for _, r in sc.iterrows()}

    b, t = pd.Timestamp("2026-04-01"), pd.Timestamp("2026-04-07")
    f0, f1 = float(fx.loc[b]), float(fx.loc[t])
    tot = 0.0
    for _, r in sc.iterrows():
        px = st[st.security_id == r.security_id].set_index("market_date")["raw_close"]
        p0, p1 = float(px.loc[b]), float(px.loc[t])
        if r.market == "US":
            p0, p1 = p0 * f0, p1 * f1
        tot += wt[r.security_id] * (p1 / p0)

    px = st[st.security_id == "010120"].set_index("market_date")["raw_close"]
    q0, q1 = float(px.loc[pd.Timestamp("2026-04-07")]), float(px.loc[pd.Timestamp("2026-04-13")])

    return {
        "A": cut.strftime("%Y-%m-%d"), "B": adtv["KTOS"], "C": 90.0, "D": p10, "E": "제외",
        "F": float((v < p10).sum()), "G": 11.0, "H": 1.0, "I": max(wt.values()) * 100,
        "J": tot * 1000, "K": (q1 / q0 - 1) * 100, "L": (q1 / (q0 / 5) - 1) * 100,
    }


def codes(path, sheet, code_col, val_col, data_only=True):
    s = load_workbook(path, data_only=data_only)[sheet]
    out = {}
    for r in range(1, s.max_row + 1):
        a = s.cell(r, code_col).value
        if isinstance(a, str) and len(a) == 1 and a.isalpha():
            out[a] = s.cell(r, val_col).value
    return out


def same(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) < max(1e-6, abs(float(a)) * 1e-9)
    return str(a).strip() == str(b).strip()


def main():
    T = truth()

    print("1) 블라인드 — 수기 워크시트에 코드 산출값이 없는가")
    wb = load_workbook(WS)
    banned = {"KTOS ADTV90": T["B"], "P10": T["D"], "지수레벨": T["J"],
              "1/6": 1 / 6, "1/18": 1 / 18}
    leaks = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)):
                    for k, val in banned.items():
                        if val and abs(c.value - val) < abs(val) * 1e-9:
                            leaks.append(f"{ws.title}!{c.coordinate}={k}")
    check("코드 산출값 누수 없음", not leaks, f"{len(leaks)}건" if leaks else "")
    for x in leaks[:10]:
        print("        ★", x)

    print("\n2) 항목 정합 — 세 파일의 대조 항목이 같은가")
    A = set(codes(WS, "09 대조표", 1, 3, data_only=False))
    B = set(codes(SEALED, "코드값(봉인)", 1, 3, data_only=False))
    C = set(codes(XLS, "99 대조표", 1, 2, data_only=False))
    check("워크시트 12항목", len(A) == 12, f"{len(A)}개")
    check("봉인 = 워크시트", A == B, f"차집합 {sorted(A ^ B)}" if A != B else "")
    check("엑셀본 = 워크시트", A == C, f"차집합 {sorted(A ^ C)}" if A != C else "")

    print("\n3) 가이드 셀 주소 — 인용한 칸이 실제 입력칸인가")
    SH = {"02": "02 자료마감일", "03-04": "03-04 거래대금·ADTV90", "05": "05 P10·편입판정",
          "06": "06 목표비중", "07": "07 지수레벨", "08": "08 분할효과"}
    g = open(GUIDE, encoding="utf-8").read()
    cur, bad, okn = None, [], 0
    for line in g.split("\n"):
        m = re.match(r"^## \d+\. (\S+)", line)
        if m and m.group(1) in SH:
            cur = SH[m.group(1)]
        m2 = re.match(r"^\| \*\*([A-Z]{1,2}\d+)(?::[A-Z]{1,2}\d+)?\*\* \| ([^|]+)\|", line)
        if m2 and cur:
            c = wb[cur][m2.group(1)]
            if c.fill.fgColor.rgb in ("00FFF7D1", "00FFE9A8"):
                okn += 1
            else:
                bad.append(f"{cur}!{m2.group(1)} ({m2.group(2).strip()})")
    check(f"가이드 인용 셀 {okn}건", not bad, f"{len(bad)}건 이상" if bad else "")
    for x in bad[:10]:
        print("        ★", x)

    print("\n4) 값 일치 — 독립 재계산 · 봉인 · 엑셀")
    S = codes(SEALED, "코드값(봉인)", 1, 5)
    X = codes(XLS, "99 대조표", 1, 3)
    mism = []
    for k in "ABCDEFGHIJKL":
        if not (same(T[k], S.get(k)) and same(T[k], X.get(k))):
            mism.append(f"{k}: 재계산={T[k]} 봉인={S.get(k)} 엑셀={X.get(k)}")
    check("12항목 삼자 일치", not mism, f"{len(mism)}건 불일치" if mism else "12/12")
    for x in mism:
        print("        ★", x)

    print("\n5) 수식 오류 — 엑셀본에 오류값이 없는가")
    xv = load_workbook(XLS, data_only=True)
    errs = [f"{s.title}!{c.coordinate}={c.value}"
            for s in xv.worksheets for row in s.iter_rows() for c in row
            if isinstance(c.value, str) and c.value in ERRS]
    check("수식 오류 없음", not errs, f"{len(errs)}건" if errs else "")
    for x in errs[:10]:
        print("        ★", x)
    blank = sum(1 for s in xv.worksheets for row in s.iter_rows() for c in row
                if c.value is None and c.fill.fgColor.rgb == "00E3EEF9")
    check("계산 칸 캐시값 존재", blank == 0,
          f"빈 계산칸 {blank}개 — recalc_with_excel.py 를 돌릴 것" if blank else "")

    print("\n" + "=" * 60)
    if fails:
        print(f"감사 실패 — {len(fails)}건: {', '.join(fails)}")
        return 1
    print("감사 통과 — 5종 검사 전부 이상 없음")
    return 0


if __name__ == "__main__":
    sys.exit(main())

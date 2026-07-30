# -*- coding: utf-8 -*-
"""
엑셀로 워크북을 재계산해 결과값을 파일에 심는다.

왜 필요한가
  openpyxl은 수식만 쓰고 계산 결과(캐시값)는 남기지 않는다. 실제 엑셀에서 열면
  자동 계산되지만, 미리보기·뷰어·pandas 로 열면 빈칸으로 보인다.
  엑셀로 한 번 열어 계산시킨 뒤 저장하면 수식과 결과값이 함께 저장돼 어디서든 보인다.

덤으로 얻는 것
  수식 문법 검증. #NAME? · #REF! · #VALUE! 가 생기면 그 자리에서 잡힌다.
  그리고 엑셀이 실제로 낸 값을 파이썬 엔진 값과 대조할 수 있다 — 이것이 교차검산의 본체다.

실행
  python recalc_with_excel.py                       # 기본 대상
  python recalc_with_excel.py 파일.xlsx             # 특정 파일
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT = os.path.join(HERE, "J5_엑셀수식_교차검산.xlsx")
ERRS = ("#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def recalc(path: str) -> None:
    import win32com.client as win32
    app = None
    try:
        app = win32.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(os.path.abspath(path))
        app.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
        print(f"[재계산] {os.path.basename(path)} — 저장 완료")
    finally:
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass


def report(path: str) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    bad = []
    filled = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERRS:
                    bad.append(f"{ws.title}!{c.coordinate} = {c.value}")
                elif c.value is not None:
                    filled += 1
    print(f"[검사] 값이 들어간 셀 {filled:,}개")
    if bad:
        print(f"[오류] 수식 오류 {len(bad)}건")
        for x in bad[:25]:
            print("   ", x)
    else:
        print("[검사] 수식 오류 없음")

    ws = wb["99 대조표"]
    print("\n[대조] 엑셀 계산값 vs 코드값")
    print(f"  {'항목':32s}{'엑셀':>22s}{'코드':>22s}  판정")
    print("  " + "-" * 80)
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not (isinstance(code, str) and len(code) == 1):
            continue
        name = str(ws.cell(r, 2).value)[:30]
        xl, cd, verdict = ws.cell(r, 3).value, ws.cell(r, 5).value, ws.cell(r, 7).value
        fx = f"{xl:,.6f}" if isinstance(xl, (int, float)) else str(xl)
        fc = f"{cd:,.6f}" if isinstance(cd, (int, float)) else str(cd)
        print(f"  {name:32s}{fx:>22s}{fc:>22s}  {verdict}")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    if not os.path.exists(target):
        raise SystemExit(f"[중단] 파일 없음: {target}")
    recalc(target)
    report(target)

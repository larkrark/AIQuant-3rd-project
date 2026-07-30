# -*- coding: utf-8 -*-
"""
J-5 수기 검산 워크북 생성기 v2 — 발표 활용 겸용

무엇인가
  코드가 낸 지수를 사람이 원자료부터 손으로 다시 계산해 대조하는 도구다.
  룰북 §14 "수기·코드 교차검산" · §18.4가 요구하는 절차다.

왜 필요한가
  독립 재산출(qa/independent)은 사람이 아니라 코드가 코드를 검증한다.
  둘 다 같은 오해를 했다면 둘 다 같은 값을 낸다. 손으로 한 번 따라가야
  그 가능성이 걸러진다.

블라인드 구조 (반드시 지킬 것)
  코드 산출값을 워크시트에 나란히 두면 답을 보며 채우게 되어 검산이 성립하지 않는다.
  그래서 두 파일로 나눈다.

    J5_수기검산_워크시트.xlsx   사람이 채우는 본. 코드 산출값이 하나도 없다.
    J5_코드값_봉인.xlsx          대조용 정답지. 수기 칸을 다 채우고 저장한 뒤에만 연다.

  ※ 이 분리가 "블라인드"를 성립시킨다. 독립 재산출은 산출물 사전 열람이 있었으므로
     블라인드가 아니라 SPEC_ASSISTED다. 두 검증의 성격이 다르다.

기준
  유니버스   Seed 18 제한 파일럿 (A안)
  산출물     data/pilot_run/output_f1
  기업행사   010120 액면분할 5:1, 조정 경계일 2026-04-13 (DART 20260205800571)

실행
  python build_worksheet_v2.py
"""
import hashlib
import os
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
import paths as P  # noqa: E402

P.force_utf8_stdout()
OUT_DIR = HERE
PILOT = os.path.join(P.DATA, "pilot_run")
PUB = os.path.join(PILOT, "output_f1")
INP = os.path.join(PILOT, "input_krxbm")

# ── 검산 대상 고정값 ──────────────────────────────────────────────
CYCLE = "2026-06-30"            # 검산 회차 (경계사례 KTOS가 여기서 탈락)
CUTOFF = "2026-06-23"           # 자료마감일 — 수기로 재도출할 대상
SAMPLE_US = "KTOS"              # ADTV90 검산 대상 (RECONSTRUCTED · 분포 최솟값)
INDEX_BASE = "2026-04-01"       # 지수 기준일
INDEX_TEST = "2026-04-07"       # 지수 검산일 (분할 이전 · 기본 산식 확인)
SPLIT_SID = "010120"
SPLIT_RATIO = 5.0
SPLIT_EFF = "2026-04-13"

# ── 서식 ─────────────────────────────────────────────────────────
NAVY = "1E2761"
C_HEAD = PatternFill("solid", fgColor=NAVY)        # 표 머리행
C_GIVEN = PatternFill("solid", fgColor="EFEFE9")   # 주어진 자료 — 수정 금지
C_INPUT = PatternFill("solid", fgColor="FFF7D1")   # 사람이 채울 칸
C_ANSWER = PatternFill("solid", fgColor="FFE9A8")  # 최종 답 칸
C_WARN = PatternFill("solid", fgColor="FDE8E6")    # 주의
C_OK = PatternFill("solid", fgColor="E6F4EA")      # 확인 완료
C_BAND = PatternFill("solid", fgColor="F7F7F4")    # 설명 영역

F_HEAD = Font(color="FFFFFF", bold=True, size=10)
F_TITLE = Font(bold=True, size=16, color=NAVY)
F_SUB = Font(size=10, color="5A5A55")
F_SEC = Font(bold=True, size=11, color=NAVY)
F_NOTE = Font(size=9, color="6B6B66")
F_BIG = Font(bold=True, size=11)

_side = Side("thin", color="C8C7BC")
BOX = Border(left=_side, right=_side, top=_side, bottom=_side)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


# ── 레이아웃 헬퍼 ────────────────────────────────────────────────
def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def title(ws, row, text, sub=None):
    ws.cell(row, 1, text).font = F_TITLE
    ws.row_dimensions[row].height = 24
    row += 1
    if sub:
        ws.cell(row, 1, sub).font = F_SUB
        row += 1
    return row + 1


def section(ws, row, text):
    c = ws.cell(row, 1, text)
    c.font = F_SEC
    return row + 1


def note(ws, row, lines, ncol=8, fill=C_BAND):
    """설명 블록 — 여러 줄을 병합 셀에 넣는다."""
    if isinstance(lines, str):
        lines = [lines]
    for ln in lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        c = ws.cell(row, 1, ln)
        c.font = F_NOTE if ln.startswith(("※", "→", "  ")) else Font(size=10)
        c.fill = fill
        c.alignment = WRAP
        row += 1
    return row + 1


def table_head(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row, start_col + i, h)
        c.fill, c.font, c.border, c.alignment = C_HEAD, F_HEAD, BOX, CTR
    return row + 1


def put(ws, row, col, val, fill=None, fmt=None, bold=False, align=None):
    c = ws.cell(row, col, val)
    c.border = BOX
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = F_BIG
    if align:
        c.alignment = align
    return c


def input_cell(ws, row, col, fmt=None, answer=False):
    return put(ws, row, col, None, C_ANSWER if answer else C_INPUT, fmt)


def verdict_validation(ws, cells):
    dv = DataValidation(type="list", formula1='"일치,불일치,미확인"', allow_blank=True)
    ws.add_data_validation(dv)
    for c in cells:
        dv.add(c)


# ── 데이터 적재 ──────────────────────────────────────────────────
def load():
    d = {}
    d["state"] = pd.read_csv(os.path.join(PUB, "daily_market_state.csv"),
                             dtype={"security_id": str}, parse_dates=["market_date"])
    d["ledger"] = pd.read_csv(os.path.join(PUB, "adtv90_ledger.csv"), dtype={"security_id": str})
    d["cons"] = pd.read_csv(os.path.join(PUB, f"constituents_{CYCLE}.csv"), dtype={"security_id": str})
    # 지수 산출구간(04-01~06-30)에 실제 적용된 것은 3/31 회차 선정분이다.
    # 6/30 회차는 선정만 되었고 적용일이 구간 밖이라 지수에 반영되지 않았다.
    d["cons_applied"] = pd.read_csv(os.path.join(PUB, "constituents_2026-03-31.csv"),
                                    dtype={"security_id": str})
    d["w"] = pd.read_csv(os.path.join(PUB, "weights_2026-03-31.csv"), dtype={"security_id": str})
    d["idx"] = pd.read_csv(os.path.join(PUB, "index_vs_benchmark.csv"), parse_dates=["market_date"])
    d["cal"] = pd.read_csv(os.path.join(INP, "calendar.csv"), parse_dates=["market_date"])
    d["fx"] = pd.read_csv(os.path.join(INP, "fx.csv"), parse_dates=["market_date"])
    return d


def common_open(cal):
    kr = set(cal[(cal.market == "KR") & (cal.is_market_open == 1)].market_date)
    us = set(cal[(cal.market == "US") & (cal.is_market_open == 1)].market_date)
    return sorted(kr & us)


# ══════════════════════════════════════════════════════════════════
# 시트 00 — 시작하기
# ══════════════════════════════════════════════════════════════════
def sheet_start(wb, meta):
    ws = wb.create_sheet("00 시작하기")
    widths(ws, {"A": 22, "B": 30, "C": 24, "D": 20, "E": 18, "F": 16, "G": 14, "H": 14})
    r = title(ws, 1, "J-5 수기 검산 워크북",
              "코드가 낸 지수를 사람이 원자료부터 손으로 다시 계산해 대조한다 · 룰북 §14 · §18.4")

    r = note(ws, r, [
        "■ 이 문서가 하는 일",
        "  코드가 계산한 값을 사람이 독립적으로 다시 계산합니다. 두 값이 같으면 구현이 규칙대로 작동한다는 증거가 됩니다.",
        "  다르면 어느 쪽이 틀렸는지 찾습니다. 어느 쪽이든 알아내는 것이 목적입니다.",
        "",
        "■ 왜 필요한가",
        "  이미 코드로 독립 재산출을 해서 10개 항목이 일치했습니다. 그러나 그것은 코드가 코드를 검증한 것입니다.",
        "  두 구현이 같은 규칙을 같이 잘못 이해했다면 둘 다 같은 값을 냅니다. 사람이 손으로 한 번 따라가야 그 경우가 걸러집니다.",
    ])

    r = note(ws, r, [
        "■ 반드시 지킬 것 — 답을 먼저 보지 않습니다",
        "  이 파일에는 코드가 낸 값이 하나도 들어 있지 않습니다. 일부러 뺐습니다.",
        "  답을 옆에 두고 채우면 무의식적으로 맞춰 쓰게 되어(확증편향) 검산이 성립하지 않습니다.",
        "  코드값은 J5_코드값_봉인.xlsx 에 따로 있습니다. 노란 칸을 전부 채우고 저장한 다음에 여십시오.",
        "  → 순서: ① 노란 칸 전부 채움  ② 파일 저장  ③ 봉인 파일 열기  ④ 09 대조표에 옮겨 적기",
    ], fill=C_WARN)

    r = section(ws, r, "색 범례")
    r = table_head(ws, r, ["색", "뜻", "다루는 법"])
    for fill, k, v in [
        (C_GIVEN, "주어진 자료", "원자료입니다. 고치지 마십시오."),
        (C_INPUT, "계산 칸", "직접 채웁니다. 엑셀 수식을 써도 됩니다."),
        (C_ANSWER, "최종 답 칸", "그 시트의 결론입니다. 대조표로 옮깁니다."),
        (C_WARN, "주의", "틀리기 쉬운 지점입니다. 읽고 진행하십시오."),
    ]:
        put(ws, r, 1, "", fill)
        put(ws, r, 2, k)
        put(ws, r, 3, v)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 엑셀 수식을 써도 되나요",
        "  됩니다. 금지되는 것은 우리 프로젝트의 파이썬 코드를 돌리는 것입니다.",
        "  주어진 원자료에서 엑셀로 평균·정렬·보간을 하는 것은 독립적인 재계산이므로 검산에 해당합니다.",
    ])

    r = section(ws, r, "검산 기준 — 이 워크북이 무엇을 대상으로 하는지")
    r = table_head(ws, r, ["항목", "값", "근거"])
    for k, v, src in [
        ("유니버스", "Seed 18 제한 파일럿 (A안)", "회의 결정 2026-07-28"),
        ("검산 회차", CYCLE, f"자료마감일 {CUTOFF}"),
        ("대상 산출물", "data/pilot_run/output_f1", "F-1 반영본"),
        ("기업행사", f"{SPLIT_SID} 액면분할 5:1", "DART 20260205800571"),
        ("조정 경계일", SPLIT_EFF + " (신주권상장일)", "효력발생일 2026-04-10 아님"),
        ("지수 기준일 / 기준값", f"{INDEX_BASE} / 1,000", "D-08"),
    ]:
        put(ws, r, 1, k, C_GIVEN)
        put(ws, r, 2, v, C_GIVEN)
        put(ws, r, 3, src, C_GIVEN)
        r += 1
    r += 1

    r = section(ws, r, "입력 스냅샷 지문 — 이 워크북이 어느 파일을 보고 만들어졌는지")
    r = table_head(ws, r, ["파일", "SHA-256 (앞 16자리)", "행 수"])
    for f, h, n in meta["hashes"]:
        put(ws, r, 1, f, C_GIVEN)
        put(ws, r, 2, h, C_GIVEN)
        put(ws, r, 3, n, C_GIVEN)
        r += 1
    r = note(ws, r, ["※ 나중에 산출물이 바뀌면 이 지문이 달라집니다. 검산 결과가 어느 버전에 대한 것인지 이것으로 특정합니다."])

    r = section(ws, r, "작성자 기록")
    for k in ["작성자", "시작 일시", "완료 일시", "봉인 개봉 일시", "사용한 도구(엑셀/계산기 등)"]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        put(ws, r, 2, None, C_INPUT)
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 01 — 검산 지도
# ══════════════════════════════════════════════════════════════════
def sheet_map(wb):
    ws = wb.create_sheet("01 검산지도")
    widths(ws, {"A": 8, "B": 26, "C": 46, "D": 26, "E": 34})
    r = title(ws, 1, "무엇을 어떤 순서로 검산하는가",
              "각 시트는 앞 시트의 결과를 이어받는다. 순서대로 진행하십시오.")

    r = table_head(ws, r, ["시트", "검산 대상", "사람이 하는 일", "근거 조항", "왜 이걸 고르는가"])
    rows = [
        ("02", "자료마감일", "선정일에서 공통 개장일 5거래일을 역산",
         "룰북 §5.2 · D-13 ⑧", "미래 정보 차단(PIT)의 출발점. 여기가 틀리면 전부 틀린다."),
        ("03", "일별 거래대금", "미국 종목의 종가×거래량 재구성",
         "룰북 제59조 · D-13 ②", "한국은 거래소가 주고 미국은 우리가 만든다. 만드는 쪽을 검산한다."),
        ("04", "ADTV90", "90개장일 평균. 정지일은 0으로 넣고 분모는 90",
         "룰북 §8.1 · D-13 ①", "결측과 무거래를 같은 0으로 합치지 않는지 확인한다."),
        ("05", "P10 · 편입판정", "8개 값을 정렬하고 선형보간으로 10퍼센타일",
         "룰북 §8.1 · §9", "경계사례가 실제로 갈리는 지점. 구조 문제도 여기서 보인다."),
        ("06", "목표비중", "3테마×2지역 6셀에 1/6씩, 셀 안은 균등",
         "D-04 ② · D-10 ①②", "합이 정확히 1이 되는지, 셀 집중도가 어떻게 되는지."),
        ("07", "지수 레벨", "기준일 대비 가중 수익률 합",
         "룰북 §13 · D-08", "최종 산출물. 여기까지 맞으면 파이프라인 전체가 맞다."),
        ("08", "분할 효과", "010120을 5:1 반영했을 때와 아닐 때의 차이",
         "데이터사전 4.1 · D-07", "성과 숫자가 왜 6.66%p 틀렸는지 직접 확인한다."),
    ]
    for a, b, c, d, e in rows:
        put(ws, r, 1, a, C_GIVEN, align=CTR)
        put(ws, r, 2, b, C_GIVEN)
        put(ws, r, 3, c, C_GIVEN).alignment = WRAP
        put(ws, r, 4, d, C_GIVEN)
        put(ws, r, 5, e, C_GIVEN).alignment = WRAP
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 표본을 이렇게 고른 이유 (룰북 §18.2 — 경계사례를 포함할 것)",
        f"  ADTV90 대상 {SAMPleFIX}: 미국 종목이라 거래대금을 우리가 재구성한다. 그리고 분포의 최솟값이라 편입 판정이 실제로 갈린다.",
        f"  기업행사 {SPLIT_SID}: 파일럿 구간의 유일한 기업행사이고, 거래정지 3일과 분할이 겹쳐 있다.",
        f"  회차 {CYCLE}: 3/31 회차는 전 종목이 여유롭게 통과해 경계가 없다. 6/30이 검산 가치가 있다.",
    ], ncol=5)

    ws.sheet_view.showGridLines = False
    return ws


SAMPleFIX = SAMPLE_US  # 문자열 포맷 편의


# ══════════════════════════════════════════════════════════════════
# 시트 02 — 자료마감일
# ══════════════════════════════════════════════════════════════════
def sheet_cutoff(wb, d):
    ws = wb.create_sheet("02 자료마감일")
    widths(ws, {"A": 6, "B": 16, "C": 12, "D": 12, "E": 16, "F": 14, "G": 30})
    r = title(ws, 1, "자료마감일 — 선정일 이전 제5거래일",
              "룰북 §5.2 · D-13 ⑧ · 축은 한·미 공통 개장일")

    r = note(ws, r, [
        "■ 규칙",
        "  선정일 당일은 세지 않습니다. 선정일 바로 앞 개장일이 1거래일 전입니다.",
        "  한국과 미국이 모두 열린 날만 셉니다. 한쪽만 열린 날은 건너뜁니다.",
        "",
        f"■ 할 일 — 선정일 {CYCLE} 에서 5거래일 앞으로 세어 자료마감일을 구하십시오.",
    ], ncol=7)

    days = [x for x in common_open(d["cal"]) if x <= pd.Timestamp(CYCLE)]
    tail = days[-9:]
    kr_open = set(d["cal"][(d["cal"].market == "KR") & (d["cal"].is_market_open == 1)].market_date)
    us_open = set(d["cal"][(d["cal"].market == "US") & (d["cal"].is_market_open == 1)].market_date)

    r = section(ws, r, "달력 — 선정일 부근 (공통 개장일만 추린 것이 아니라 원본 그대로입니다)")
    r = table_head(ws, r, ["", "날짜", "한국 개장", "미국 개장", "공통 개장일?", "며칠 전?", "메모"])
    alld = sorted(set(list(kr_open) + list(us_open)))
    win = [x for x in alld if pd.Timestamp("2026-06-15") <= x <= pd.Timestamp(CYCLE)]
    first_row = r
    for i, dt in enumerate(win):
        put(ws, r, 1, i + 1, C_GIVEN, align=CTR)
        put(ws, r, 2, dt.strftime("%Y-%m-%d"), C_GIVEN)
        put(ws, r, 3, "O" if dt in kr_open else "-", C_GIVEN, align=CTR)
        put(ws, r, 4, "O" if dt in us_open else "-", C_GIVEN, align=CTR)
        input_cell(ws, r, 5).alignment = CTR      # 공통 개장일 여부
        input_cell(ws, r, 6).alignment = CTR      # 며칠 전
        put(ws, r, 7, None, C_INPUT)
        r += 1
    r += 1

    r = section(ws, r, "답")
    put(ws, r, 1, "자료마감일", C_GIVEN)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    input_cell(ws, r, 2, answer=True)
    put(ws, r, 4, "← 이 값을 09 대조표로 옮깁니다", C_BAND)
    r += 2

    r = note(ws, r, [
        "■ 틀리기 쉬운 곳",
        "  선정일을 1거래일 전으로 세면 하루 밀립니다. 선정일 다음이 아니라 그 앞이 1거래일 전입니다.",
        "  한쪽 시장만 열린 날을 세면 결과가 달라집니다. 실제로 이 구간에 그런 날이 있는지 확인해 보십시오.",
    ], ncol=7, fill=C_WARN)

    ws.freeze_panes = ws.cell(first_row, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 03·04 — 거래대금 재구성 + ADTV90
# ══════════════════════════════════════════════════════════════════
def sheet_adtv(wb, d):
    ws = wb.create_sheet("03-04 거래대금·ADTV90")
    widths(ws, {"A": 6, "B": 14, "C": 14, "D": 16, "E": 18, "F": 16, "G": 26})
    r = title(ws, 1, f"{SAMPLE_US} 의 ADTV90 — 90개장일 평균 거래대금",
              "룰북 제59조(거래대금 원천) · §8.1(ADTV90) · D-13 ①(정지일 0 반영)")

    r = note(ws, r, [
        "■ 규칙 두 가지",
        "  ① 거래대금: 한국은 거래소가 제공한 값을 그대로 씁니다. 미국은 제공값이 없어 종가 × 거래량으로 재구성합니다.",
        f"     {SAMPLE_US}는 미국 종목이므로 재구성 대상입니다. D열을 직접 계산하십시오.",
        "  ② 평균: 관측창은 자료마감일부터 뒤로 90 개장일입니다. 거래정지일은 0으로 넣고 분모에서 빼지 않습니다(분모 90).",
        "     결측(자료를 못 구한 날)은 다릅니다. 결측은 분모에서 뺍니다. 둘을 같은 0으로 합치면 안 됩니다.",
    ], ncol=7)

    st = d["state"]
    us_open = sorted(d["cal"][(d["cal"].market == "US") & (d["cal"].is_market_open == 1)].market_date)
    win = [x for x in us_open if x <= pd.Timestamp(CUTOFF)][-90:]
    sub = st[(st.security_id == SAMPLE_US) & (st.market_date.isin(win))].sort_values("market_date")

    r = section(ws, r, f"원자료 — {win[0].date()} ~ {win[-1].date()} (미국 개장일 90일)")
    r = table_head(ws, r, ["", "날짜", "종가(USD)", "거래량(주)", "거래대금(직접 계산)", "상태코드", "메모"])
    head = r
    for i, (_, row) in enumerate(sub.iterrows()):
        put(ws, r, 1, i + 1, C_GIVEN, align=CTR)
        put(ws, r, 2, row.market_date.strftime("%Y-%m-%d"), C_GIVEN)
        put(ws, r, 3, float(row.raw_close), C_GIVEN, "#,##0.00")
        put(ws, r, 4, float(row.volume) if pd.notna(row.volume) else None, C_GIVEN, "#,##0")
        input_cell(ws, r, 5, "#,##0")
        put(ws, r, 6, row.daily_market_state, C_GIVEN, align=CTR)
        put(ws, r, 7, None, C_INPUT)
        r += 1
    last = r - 1
    r += 1

    r = section(ws, r, "집계")
    for k, hint in [
        ("관측 개장일수 (분모가 되는 수)", "위 표의 행 수"),
        ("그중 거래정지일 수", "상태코드가 TRADING_HALT 인 날"),
        ("그중 결측일 수", "자료 자체가 없는 날 — 분모에서 뺄 대상"),
        ("거래대금 합계", f"E{head}:E{last} 의 합"),
        ("ADTV90 = 합계 ÷ 분모", "분모를 무엇으로 잡았는지 메모에 적으십시오"),
    ]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        input_cell(ws, r, 2, "#,##0.00", answer=("ADTV90" in k))
        put(ws, r, 4, hint, C_BAND)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 스스로 점검할 것",
        "  분모를 90으로 했습니까, 아니면 정지일을 뺀 수로 했습니까? 규칙은 90입니다. 왜 그런지 설명할 수 있습니까?",
        "  만약 정지일을 분모에서 뺐다면 ADTV90이 커집니다. 편입 판정이 뒤집힐 수 있습니다. 다음 시트에서 확인됩니다.",
    ], ncol=7, fill=C_WARN)

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 05 — P10 · 편입판정
# ══════════════════════════════════════════════════════════════════
def sheet_p10(wb, d):
    ws = wb.create_sheet("05 P10·편입판정")
    widths(ws, {"A": 6, "B": 14, "C": 22, "D": 16, "E": 18, "F": 18, "G": 30})
    r = title(ws, 1, "P10 하한과 편입 판정",
              "룰북 §8.1(분포·하한) · §9(편입) · 선형보간")

    r = note(ws, r, [
        "■ 규칙",
        "  같은 시장 안에서 ADTV90이 산출된 종목들을 모아 10퍼센타일을 구합니다. 그보다 낮으면 유동성 미달로 제외합니다.",
        "  시즈닝(상장 후 유효관측 90일)을 못 채운 종목은 ADTV90 자체가 산출되지 않아 분포에서 빠집니다.",
        "",
        "■ 선형보간으로 퍼센타일 구하는 법",
        "  값을 작은 순으로 정렬하고 0부터 번호를 붙입니다. 위치 h = (n − 1) × (백분위 ÷ 100).",
        "  h가 정수가 아니면 아래위 두 값 사이를 비례로 나눕니다.  값 = x[내림] + (h − 내림) × (x[올림] − x[내림])",
        "",
        f"■ 할 일 — 아래 미국 종목 분포에서 P10을 구하고 {SAMPLE_US}의 편입 여부를 판정하십시오.",
        f"  {SAMPLE_US} 칸은 비어 있습니다. 앞 시트에서 직접 계산한 값을 넣으십시오.",
    ], ncol=7)

    led = d["ledger"]
    g = led[(led.market == "US") & (led.selection_date == CYCLE)]
    calc = g.dropna(subset=["official_adtv90"]).copy()
    excl = g[g.official_adtv90.isna()]

    r = section(ws, r, "미국 시장 ADTV90 분포 — 검산 대상 종목만 비워 두었습니다")
    r = table_head(ws, r, ["", "종목", "ADTV90 (USD)", "시즈닝", "분포 포함?", "정렬 후 순번", "메모"])
    head = r
    for _, row in calc.sort_values("security_id").iterrows():
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, row.security_id, C_GIVEN, align=CTR)
        if row.security_id == SAMPLE_US:
            input_cell(ws, r, 3, "#,##0.00", answer=True)
            put(ws, r, 7, "← 03-04 시트에서 직접 계산한 값", C_WARN)
        else:
            put(ws, r, 3, float(row.official_adtv90), C_GIVEN, "#,##0.00")
            put(ws, r, 7, None, C_INPUT)
        put(ws, r, 4, row.seasoning_status, C_GIVEN, align=CTR)
        input_cell(ws, r, 5).alignment = CTR
        input_cell(ws, r, 6).alignment = CTR
        r += 1
    for _, row in excl.iterrows():
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, row.security_id, C_GIVEN, align=CTR)
        put(ws, r, 3, "(산출 안 됨)", C_GIVEN, align=CTR)
        put(ws, r, 4, row.seasoning_status, C_GIVEN, align=CTR)
        input_cell(ws, r, 5).alignment = CTR
        input_cell(ws, r, 6).alignment = CTR
        put(ws, r, 7, "분포에 넣을지 판단하십시오", C_INPUT)
        r += 1
    r += 1

    r = section(ws, r, "계산")
    for k, hint in [
        ("분포에 들어간 종목 수 n", "ADTV90이 산출된 종목만"),
        ("보간 위치 h = (n−1) × 0.10", "소수로 나옵니다"),
        ("h 아래 정수 위치의 값", "정렬 후 x[내림]"),
        ("h 위 정수 위치의 값", "정렬 후 x[올림]"),
        ("P10 하한", "위 보간식을 적용"),
        (f"{SAMPLE_US} 편입 여부", "P10 이상이면 편입, 미만이면 제외"),
        (f"{SAMPLE_US}의 P10 대비 차이(%)", "(ADTV90 ÷ P10 − 1) × 100"),
        ("P10 미만인 종목 수", "몇 개나 떨어졌습니까?"),
    ]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        input_cell(ws, r, 2, answer=("P10 하한" in k or "편입 여부" in k))
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        put(ws, r, 4, hint, C_BAND)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 다 구한 뒤에 생각해 볼 것 — 이것이 발표에서 쓸 내용입니다",
        "  h를 계산해 보면 0과 1 사이입니다. 그러면 P10은 항상 '가장 작은 값'과 '두 번째로 작은 값' 사이에 놓입니다.",
        "  즉 P10 미만인 종목은 언제나 정확히 한 개입니다. 그 종목의 거래대금이 크든 작든 상관없습니다.",
        "  → 지금 규칙은 '유동성이 일정 수준 미만이면 제외'가 아니라 '가장 낮은 하나를 제외'로 작동하고 있습니다.",
        "  → 종목 수가 몇 개가 되어야 두 개 이상 떨어질 수 있습니까? h ≥ 1 이 되는 n을 구해 보십시오.",
    ], ncol=7)
    put(ws, r, 1, "두 개 이상 떨어지려면 필요한 최소 n", C_GIVEN)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    input_cell(ws, r, 2, answer=True)

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 06 — 목표비중
# ══════════════════════════════════════════════════════════════════
def sheet_weight(wb, d):
    ws = wb.create_sheet("06 목표비중")
    widths(ws, {"A": 6, "B": 14, "C": 18, "D": 16, "E": 16, "F": 18, "G": 18, "H": 24})
    r = title(ws, 1, "6셀 구조와 목표비중",
              "D-04 ②(테마 1:1:1) · D-10 ①(지역 50:50) · D-10 ②(6셀 각 1/6)")

    r = note(ws, r, [
        "■ 규칙",
        "  테마 3개(우주방산·AI로보틱스·에너지전력)에 각각 1/3, 지역 2개(한국·미국)에 각각 1/2.",
        "  두 조건을 함께 만족하도록 3×2 = 6개 셀에 각 1/6씩 배분합니다.",
        "  셀 안에서는 편입된 종목끼리 균등하게 나눕니다. 셀에 1종목뿐이면 그 종목이 1/6을 전부 갖습니다.",
        "",
        "■ 할 일 — 각 셀의 편입 종목 수를 세고, 종목별 목표비중을 계산하십시오.",
    ], ncol=8)

    r = note(ws, r, [
        "■ 주의 — 여기서 쓰는 회차는 05 시트와 다릅니다. 헷갈리기 쉬운 지점입니다.",
        "  05 시트는 2026-06-30 회차의 편입 판정이었습니다. 그 결과는 '다음 구간'에 적용될 것입니다.",
        "  지금 지수 산출구간(2026-04-01~06-30)에 실제로 적용된 것은 2026-03-31 회차 선정분입니다.",
        "  그래서 이 시트는 3/31 회차를 씁니다. 07 시트의 지수 계산도 이 비중을 씁니다.",
        "  → 선정일과 적용일은 다릅니다. 6/30 선정분은 적용일이 이 구간 밖이라 지수에 들어가지 않았습니다.",
        "  → 참고로 룰북 245행은 적용일 세부 규칙을 아직 후속과제로 두고 있습니다. 산출물에 effective_date 열이 없습니다.",
    ], ncol=8, fill=C_WARN)

    cons = d["cons_applied"]
    sel = cons[cons.selected_flag == 1]

    r = section(ws, r, "셀별 집계")
    r = table_head(ws, r, ["", "셀", "테마", "지역", "셀 목표비중", "편입 종목수", "종목당 비중", "메모"])
    head = r
    for cell in sorted(sel.cell_id.unique()):
        g = sel[sel.cell_id == cell]
        mk, th = cell.split("_", 1)
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, cell, C_GIVEN)
        put(ws, r, 3, th, C_GIVEN)
        put(ws, r, 4, mk, C_GIVEN, align=CTR)
        input_cell(ws, r, 5, "0.000000")
        input_cell(ws, r, 6).alignment = CTR
        input_cell(ws, r, 7, "0.000000")
        put(ws, r, 8, None, C_INPUT)
        r += 1
    put(ws, r, 4, "합계", C_GIVEN, bold=True)
    input_cell(ws, r, 5, "0.000000", answer=True)
    put(ws, r, 8, "정확히 1.000000 이어야 합니다", C_WARN)
    r += 2

    r = section(ws, r, "종목별 목표비중 — 편입 종목 전체")
    r = table_head(ws, r, ["", "종목", "지역", "테마", "셀", "목표비중", "", "메모"])
    h2 = r
    for _, row in sel.sort_values(["cell_id", "security_id"]).iterrows():
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, row.security_id, C_GIVEN, align=CTR)
        put(ws, r, 3, row.market, C_GIVEN, align=CTR)
        put(ws, r, 4, row.primary_theme, C_GIVEN)
        put(ws, r, 5, row.cell_id, C_GIVEN)
        input_cell(ws, r, 6, "0.000000")
        put(ws, r, 7, None, C_GIVEN)
        put(ws, r, 8, None, C_INPUT)
        r += 1
    put(ws, r, 5, "합계", C_GIVEN, bold=True)
    input_cell(ws, r, 6, "0.000000", answer=True)
    r += 2

    r = note(ws, r, [
        "■ 다 구한 뒤에 생각해 볼 것",
        "  가장 큰 비중을 가진 종목은 몇 퍼센트입니까? 다른 종목의 몇 배입니까?",
        "  그 셀에 종목이 몇 개 있습니까? 그 종목이 다음 회차에 빠지면 그 셀은 어떻게 됩니까?",
        "  → 지금 규칙에는 개별종목 상한이 없고 셀 최소 종목수도 정해져 있지 않습니다.",
    ], ncol=8)
    for k in ["최대 비중 종목", "그 비중(%)", "그 셀의 종목 수"]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        input_cell(ws, r, 2, answer=True)
        r += 1

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 07 — 지수 레벨
# ══════════════════════════════════════════════════════════════════
def sheet_index(wb, d):
    ws = wb.create_sheet("07 지수레벨")
    widths(ws, {"A": 6, "B": 12, "C": 10, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 16, "J": 20})
    r = title(ws, 1, f"지수 레벨 — {INDEX_TEST}",
              "룰북 §13 · D-08 · 기준일 " + INDEX_BASE + " = 1,000 · 원화 · 무헤지")

    r = note(ws, r, [
        "■ 규칙",
        "  구간 시작일 가격 대비 현재 가격의 비율을 종목마다 구하고, 목표비중을 곱해 모두 더합니다.",
        "  거기에 기준값 1,000을 곱하면 지수 레벨입니다.",
        "  미국 종목은 그날의 환율로 원화로 바꾼 뒤에 비율을 계산합니다. 환율도 날마다 다릅니다.",
        "",
        "      지수 = 1000 × Σ ( 목표비중 × 당일가격(원) ÷ 기준일가격(원) )",
        "",
        f"■ 할 일 — {INDEX_TEST} 의 지수 레벨을 구하십시오. (분할 이전 날짜라 기업행사 문제는 없습니다)",
        "  목표비중(D열)은 주어져 있지 않습니다. 06 시트에서 직접 계산한 값을 옮겨 적으십시오.",
    ], ncol=10)

    st, w, fx = d["state"], d["w"], d["fx"]
    fxs = fx.set_index("market_date")["fx_rate"]
    b, t = pd.Timestamp(INDEX_BASE), pd.Timestamp(INDEX_TEST)
    fb, ft = float(fxs.loc[b]), float(fxs.loc[t])

    r = section(ws, r, "환율 (ECOS 매매기준율, 평가일 당일 적용)")
    put(ws, r, 1, f"기준일 {INDEX_BASE}", C_GIVEN)
    put(ws, r, 3, fb, C_GIVEN, "#,##0.0")
    put(ws, r, 5, f"검산일 {INDEX_TEST}", C_GIVEN)
    put(ws, r, 7, ft, C_GIVEN, "#,##0.0")
    r += 2

    r = table_head(ws, r, ["", "종목", "지역", "목표비중",
                           f"기준일 종가", f"검산일 종가",
                           "기준일 가격(원)", "검산일 가격(원)", "가격비", "비중×가격비"])
    head = r
    for _, row in w.sort_values(["market", "security_id"]).iterrows():
        sid = row.security_id
        px = st[st.security_id == sid].set_index("market_date")["raw_close"]
        p0 = float(px.loc[b]) if b in px.index else None
        p1 = float(px.loc[t]) if t in px.index else None
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, sid, C_GIVEN, align=CTR)
        put(ws, r, 3, row.market, C_GIVEN, align=CTR)
        # 비중은 주지 않는다 — 주면 06 시트의 답이 새서 블라인드가 깨진다.
        input_cell(ws, r, 4, "0.000000")
        put(ws, r, 5, p0, C_GIVEN, "#,##0.00")
        put(ws, r, 6, p1, C_GIVEN, "#,##0.00")
        input_cell(ws, r, 7, "#,##0.00")
        input_cell(ws, r, 8, "#,##0.00")
        input_cell(ws, r, 9, "0.000000")
        input_cell(ws, r, 10, "0.000000")
        r += 1
    last = r - 1
    put(ws, r, 9, "합계", C_GIVEN, bold=True)
    input_cell(ws, r, 10, "0.000000")
    r += 2

    r = section(ws, r, "답")
    put(ws, r, 1, f"{INDEX_TEST} 지수 레벨 = 1000 × 합계", C_GIVEN)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    input_cell(ws, r, 4, "#,##0.0000", answer=True)
    r += 2

    r = note(ws, r, [
        "■ 틀리기 쉬운 곳",
        "  한국 종목에 환율을 곱하면 안 됩니다. 이미 원화입니다.",
        "  미국 종목은 기준일과 검산일에 서로 다른 환율을 씁니다. 같은 환율을 쓰면 환율 변동분이 사라집니다.",
        "  비중은 소수점 아래가 깁니다(1/18 = 0.055556). 반올림하면 합계가 어긋납니다.",
    ], ncol=10, fill=C_WARN)

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 08 — 분할 효과
# ══════════════════════════════════════════════════════════════════
def sheet_split(wb, d):
    ws = wb.create_sheet("08 분할효과")
    widths(ws, {"A": 6, "B": 14, "C": 14, "D": 16, "E": 16, "F": 18, "G": 34})
    r = title(ws, 1, f"{SPLIT_SID} 액면분할 5:1 — 반영했을 때와 안 했을 때",
              "데이터사전 4.1 · D-07 · DART 20260205800493 / 20260205800571(정정)")

    r = note(ws, r, [
        "■ 무슨 일이 있었나",
        "  2026-02-05 이사회에서 1주를 5주로 쪼개기로 결정했습니다(액면가 5,000원 → 1,000원).",
        "  2026-04-08~10 사흘간 거래가 정지됐고 04-13에 재개됐습니다.",
        "  재산은 그대로인데 주가 숫자만 1/5이 됩니다. 이것을 반영하지 않으면 주가가 폭락한 것처럼 보입니다.",
        "",
        "■ 조정 방법 (데이터사전 4.1)",
        f"  효력일({SPLIT_EFF}) '이전' 가격을 분할비율 5로 나눕니다. 이후 가격은 그대로 둡니다.",
        "  거래정지 기간의 가격도 분할 전 기준이므로 함께 나눕니다.",
        "",
        "■ 할 일 — 아래 두 열을 채우고 04-13의 수익률을 두 가지로 구해 비교하십시오.",
    ], ncol=7)

    st = d["state"]
    win = pd.date_range("2026-04-06", "2026-04-16")
    sub = st[(st.security_id == SPLIT_SID) & (st.market_date.isin(win))].sort_values("market_date")

    r = table_head(ws, r, ["", "날짜", "원종가", "상태코드", "분할조정 종가", "전일대비 수익률", "메모"])
    head = r
    for _, row in sub.iterrows():
        put(ws, r, 1, None, C_GIVEN)
        put(ws, r, 2, row.market_date.strftime("%Y-%m-%d"), C_GIVEN)
        put(ws, r, 3, float(row.raw_close), C_GIVEN, "#,##0")
        put(ws, r, 4, row.daily_market_state, C_GIVEN, align=CTR)
        input_cell(ws, r, 5, "#,##0.0")
        input_cell(ws, r, 6, "0.00%")
        put(ws, r, 7, None, C_INPUT)
        r += 1
    r += 1

    r = section(ws, r, "04-13 수익률 비교")
    for k, hint in [
        ("① 조정 안 했을 때 (원종가 그대로)", "179,200 ÷ 788,000 − 1"),
        ("② 조정 했을 때", "179,200 ÷ (788,000÷5) − 1"),
        ("③ 차이 (① − ②)", "이만큼이 실재하지 않는 하락입니다"),
        (f"④ 지수에 준 영향 = ③ × {SPLIT_SID} 목표비중", "비중은 06 시트에서"),
    ]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        input_cell(ws, r, 3, "0.00%", answer=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        put(ws, r, 4, hint, C_BAND)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 자기 점검 — 여기가 가장 중요합니다",
        "  조정한 뒤 04-13의 수익률이 0% 가 나왔다면 두 번 나눈 것입니다(과보정). 다시 확인하십시오.",
        "  올바르게 조정하면 04-13은 '플러스' 수익률이 나옵니다. 분할 기준가보다 높게 거래가 재개됐기 때문입니다.",
        "  → 실제 시장이 분할을 호재로 봤다는 뜻이고, 그것이 진짜 정보입니다. 조정 전 숫자는 그것을 지워버립니다.",
        "",
        "■ 조정 경계일을 04-10(효력발생일)으로 잡으면 어떻게 됩니까?",
        "  공시에는 신주 효력발생일이 04-10, 신주권상장일이 04-13으로 따로 적혀 있습니다.",
        "  04-10을 경계로 잡으면 정지 기간의 가격이 조정되지 않아 왜곡이 04-13에서 04-10으로 자리만 옮깁니다.",
        "  → 가격 계열에서는 거래가 실제로 재개된 04-13이 경계입니다. 직접 확인해 보십시오.",
    ], ncol=7, fill=C_WARN)

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 09 — 대조표
# ══════════════════════════════════════════════════════════════════
ITEMS = [
    ("A", "02", "자료마감일", "날짜"),
    ("B", "03-04", f"{SAMPLE_US} ADTV90", "USD"),
    ("C", "03-04", "ADTV90 분모(개장일수)", "일"),
    ("D", "05", "미국 P10 하한", "USD"),
    ("E", "05", f"{SAMPLE_US} 편입 여부", "편입/제외"),
    ("F", "05", "P10 미만 종목 수", "개"),
    ("G", "05", "2종목 이상 탈락에 필요한 최소 n", "개"),
    ("H", "06", "목표비중 합계 (3/31 회차)", "비율"),
    ("I", "06", "최대 비중 종목의 비중 (3/31 회차)", "%"),
    ("J", "07", f"{INDEX_TEST} 지수 레벨", "포인트"),
    ("K", "08", "04-13 조정 전 수익률", "%"),
    ("L", "08", "04-13 조정 후 수익률", "%"),
]


def sheet_compare(wb):
    ws = wb.create_sheet("09 대조표")
    widths(ws, {"A": 6, "B": 8, "C": 30, "D": 12, "E": 20, "F": 20, "G": 14, "H": 40})
    r = title(ws, 1, "수기값 · 코드값 대조표",
              "노란 칸을 전부 채우고 저장한 뒤에 J5_코드값_봉인.xlsx 를 열어 F열을 채웁니다")

    r = note(ws, r, [
        "■ 순서",
        "  ① 앞 시트의 답 칸을 E열에 옮겨 적습니다.  ② 이 파일을 저장합니다.",
        "  ③ 그 다음에 봉인 파일을 열어 F열을 채웁니다.  ④ G열에 일치 여부를 고릅니다.",
        "  불일치가 나오면 반드시 H열에 원인을 적습니다. 불일치를 지우지 마십시오 — 그것이 검산의 성과입니다.",
    ], fill=C_WARN)

    r = table_head(ws, r, ["", "시트", "검산 항목", "단위", "수기값", "코드값", "판정", "불일치 원인 / 메모"])
    head = r
    vcells = []
    for code, sh, name, unit in ITEMS:
        put(ws, r, 1, code, C_GIVEN, align=CTR)
        put(ws, r, 2, sh, C_GIVEN, align=CTR)
        put(ws, r, 3, name, C_GIVEN)
        put(ws, r, 4, unit, C_GIVEN, align=CTR)
        input_cell(ws, r, 5)
        input_cell(ws, r, 6)
        c = input_cell(ws, r, 7)
        c.alignment = CTR
        vcells.append(f"G{r}")
        put(ws, r, 8, None, C_INPUT)
        r += 1
    verdict_validation(ws, vcells)
    r += 1

    r = section(ws, r, "종합")
    for k in ["일치 항목 수", "불일치 항목 수", "검산 결론 (한 문장)"]:
        put(ws, r, 1, k, C_GIVEN)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        input_cell(ws, r, 4, answer=True)
        r += 1

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 시트 10 — 발표용 요약
# ══════════════════════════════════════════════════════════════════
def sheet_present(wb):
    ws = wb.create_sheet("10 발표용요약")
    widths(ws, {"A": 4, "B": 30, "C": 26, "D": 26, "E": 12, "F": 46})
    r = title(ws, 1, "발표용 요약 — 채우고 나면 이 시트를 그대로 쓸 수 있습니다",
              "무엇을 검증했고 무엇이 확인됐는지 한 장으로 보여 주는 용도")

    r = note(ws, r, [
        "■ 이 시트의 쓰임",
        "  대조표를 채운 뒤 결과를 여기에 옮기면 발표 슬라이드 한 장이 됩니다.",
        "  성과 숫자가 아니라 '우리 숫자를 어디까지 믿을 수 있는지'를 보여 주는 자료입니다.",
    ])

    r = section(ws, r, "검증 결과")
    r = table_head(ws, r, ["", "검증한 것", "수기값", "코드값", "판정", "이 항목이 확인해 주는 것"])
    meaning = {
        "A": "미래 정보가 섞이지 않았다 (PIT)",
        "B": "미국 거래대금 재구성이 제59조대로 되었다",
        "C": "정지일과 결측일을 다르게 처리했다 (R6)",
        "D": "유동성 하한이 규칙대로 산출되었다",
        "E": "경계 종목의 판정이 임의적이지 않다",
        "F": "지금 규칙이 매 회차 정확히 1종목을 제외한다",
        "G": "그 성질이 표본 크기에서 나온 것이다",
        "H": "비중 합이 정확히 1이다",
        "I": "개별종목 집중도가 어느 수준인지",
        "J": "최종 지수값이 규칙대로 산출되었다",
        "K": "기업행사 미반영 시 얼마나 왜곡되는지",
        "L": "조정 후 값이 실제 시장 반응과 맞는지",
    }
    head = r
    for code, sh, name, unit in ITEMS:
        put(ws, r, 1, code, C_GIVEN, align=CTR)
        put(ws, r, 2, name, C_GIVEN)
        put(ws, r, 3, f"='09 대조표'!E{head + (r - head)}", C_INPUT)
        put(ws, r, 4, f"='09 대조표'!F{head + (r - head)}", C_INPUT)
        put(ws, r, 5, f"='09 대조표'!G{head + (r - head)}", C_INPUT, align=CTR)
        put(ws, r, 6, meaning[code], C_GIVEN).alignment = WRAP
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 발표에서 말할 문장 (검산 결과에 맞게 고쳐 쓰십시오)",
        "  우리는 코드가 낸 지수를 사람이 원자료부터 손으로 다시 계산해 대조했습니다.",
        "  자료마감일, 거래대금 재구성, ADTV90, 유동성 하한, 목표비중, 지수 레벨까지 12개 항목입니다.",
        "  이 과정에서 규칙 자체의 성질도 확인했습니다 — 지금의 유동성 하한은 후보군이 작아서",
        "  매 회차 가장 낮은 한 종목만 제외하는 방식으로 작동합니다. 이것은 알고 채택한 것입니다.",
        "",
        "■ 하지 말아야 할 말",
        "  '검산이 맞았으니 성과도 믿을 만하다' — 검산은 계산이 규칙대로 되었는지만 말합니다.",
        "  성과의 유의성은 표본 58일, 추적오차 33%로 통계적으로 판정할 수 없습니다. 이 둘은 다른 이야기입니다.",
    ], ncol=6, fill=C_WARN)

    ws.freeze_panes = ws.cell(head, 1)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════
# 봉인 파일
# ══════════════════════════════════════════════════════════════════
def build_sealed(d, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "코드값(봉인)"
    widths(ws, {"A": 6, "B": 8, "C": 32, "D": 12, "E": 26, "F": 50})
    r = title(ws, 1, "코드 산출값 — 봉인",
              "수기 워크시트를 전부 채우고 저장한 뒤에만 여십시오")
    r = note(ws, r, [
        "■ 먼저 읽으십시오",
        "  이 파일을 미리 열면 수기 검산이 성립하지 않습니다. 답을 본 뒤의 계산은 검증이 아닙니다.",
        "  워크시트를 다 채웠다면, 아래 값을 09 대조표 F열에 옮겨 적으십시오.",
    ], ncol=6, fill=C_WARN)

    led, cons = d["ledger"], d["cons"]
    g = led[(led.market == "US") & (led.selection_date == CYCLE)]
    calc = g.dropna(subset=["official_adtv90"])
    v = np.sort(calc.official_adtv90.values)
    n = len(v)
    p10 = float(np.percentile(v, 10))
    ktos = float(calc[calc.security_id == SAMPLE_US].official_adtv90.iloc[0])
    krow = led[(led.security_id == SAMPLE_US) & (led.selection_date == CYCLE)].iloc[0]
    sel = cons[cons.selected_flag == 1]
    wmax = d["w"].final_target_weight.max()
    wtop = d["w"].loc[d["w"].final_target_weight.idxmax(), "security_id"]
    idx = d["idx"].set_index("market_date")["index_level"]
    lvl = float(idx.loc[pd.Timestamp(INDEX_TEST)])
    st = d["state"]
    px = st[st.security_id == SPLIT_SID].set_index("market_date")["raw_close"]
    p_before, p_after = float(px.loc[pd.Timestamp("2026-04-07")]), float(px.loc[pd.Timestamp(SPLIT_EFF)])
    r_raw = p_after / p_before - 1
    r_adj = p_after / (p_before / SPLIT_RATIO) - 1

    # 값은 문자열이 아니라 숫자로 넣는다. 반올림해 두면 엑셀 교차검산본과 정밀 대조가 안 된다.
    # 단위는 09 대조표가 선언한 단위와 일치시킨다 (I·K·L 은 %).
    vals = {
        "A": (CUTOFF, "@", "룰북 §5.2 · 공통 개장일 축 5거래일 역산"),
        "B": (float(ktos), "#,##0.000000", f"official_adtv90 · 방법 {krow.official_adtv90_method}"),
        "C": (int(krow.observed_open_days), "#,##0",
              f"정지 {int(krow.halt_days_90)}일은 0으로 넣고 분모에 포함 · 결측 {int(krow.missing_days_90)}일"),
        "D": (float(p10), "#,##0.000000", f"n={n} · h=(n−1)×0.10={round((n-1)*0.10,2)} · 선형보간"),
        "E": ("제외", "@", f"ADTV90 {ktos:,.2f} < P10 {p10:,.2f}"),
        "F": (int((v < p10).sum()), "#,##0", "h<1 이므로 항상 최솟값 1개만 미달"),
        "G": (11, "#,##0", "h=(n−1)×0.10 ≥ 1 이 되는 최소 n"),
        "H": (1.0, "0.000000", "6셀 × 1/6"),
        "I": (float(wmax * 100), "0.000000", f"{wtop} · 셀에 1종목뿐이라 1/6을 단독 보유 · 단위 %"),
        "J": (float(lvl), "#,##0.000000", f"기준일 {INDEX_BASE}=1,000 · SEGMENT_RELINK · SAME_DAY_ECOS"),
        "K": (float(r_raw * 100), "0.000000",
              f"{p_after:,.0f} ÷ {p_before:,.0f} − 1 · 실재하지 않는 하락 · 단위 %"),
        "L": (float(r_adj * 100), "0.000000",
              f"{p_after:,.0f} ÷ {p_before / SPLIT_RATIO:,.0f} − 1 · 분할 기준가 대비 실제 시장 반응 · 단위 %"),
    }

    r = table_head(ws, r, ["", "시트", "검산 항목", "단위", "코드값", "산출 근거"])
    for code, sh, name, unit in ITEMS:
        val, fmt, why = vals[code]
        put(ws, r, 1, code, C_GIVEN, align=CTR)
        put(ws, r, 2, sh, C_GIVEN, align=CTR)
        put(ws, r, 3, name, C_GIVEN)
        put(ws, r, 4, unit, C_GIVEN, align=CTR)
        put(ws, r, 5, val, C_OK, fmt, bold=True)
        put(ws, r, 6, why, C_GIVEN).alignment = WRAP
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 불일치가 났을 때",
        "  코드가 맞고 수기가 틀렸을 수도, 그 반대일 수도 있습니다. 어느 쪽인지 따지는 것이 이 절차의 목적입니다.",
        "  자주 갈리는 지점: ADTV90 분모(90인가 87인가) · 환율을 한국 종목에도 곱했는가 · 분할 조정 경계일",
        "  불일치를 발견하면 지우지 말고 09 대조표 H열에 원인을 적은 뒤 팀에 공유하십시오.",
    ], ncol=6)

    ws.sheet_view.showGridLines = False
    wb.save(path)


# ══════════════════════════════════════════════════════════════════
def snapshot_hashes():
    files = [
        (os.path.join(PUB, "daily_market_state.csv"), "daily_market_state.csv"),
        (os.path.join(PUB, "adtv90_ledger.csv"), "adtv90_ledger.csv"),
        (os.path.join(PUB, f"constituents_{CYCLE}.csv"), f"constituents_{CYCLE}.csv"),
        (os.path.join(PUB, "weights_2026-03-31.csv"), "weights_2026-03-31.csv"),
        (os.path.join(PUB, "index_vs_benchmark.csv"), "index_vs_benchmark.csv"),
        (os.path.join(INP, "calendar.csv"), "calendar.csv"),
        (os.path.join(INP, "fx.csv"), "fx.csv"),
    ]
    out = []
    for full, name in files:
        with open(full, "rb") as f:
            raw = f.read()
        h = hashlib.sha256(raw).hexdigest()[:16]
        n = sum(1 for _ in open(full, encoding="utf-8-sig")) - 1
        out.append((name, h, n))
    return out


def main():
    d = load()
    meta = {"hashes": snapshot_hashes()}

    wb = Workbook()
    wb.remove(wb.active)
    sheet_start(wb, meta)
    sheet_map(wb)
    sheet_cutoff(wb, d)
    sheet_adtv(wb, d)
    sheet_p10(wb, d)
    sheet_weight(wb, d)
    sheet_index(wb, d)
    sheet_split(wb, d)
    sheet_compare(wb)
    sheet_present(wb)

    ws_path = os.path.join(OUT_DIR, "J5_수기검산_워크시트.xlsx")
    sealed_path = os.path.join(OUT_DIR, "J5_코드값_봉인.xlsx")
    wb.save(ws_path)
    build_sealed(d, sealed_path)

    pd.DataFrame(meta["hashes"], columns=["file", "sha256_16", "rows"]).to_csv(
        os.path.join(OUT_DIR, "snapshot_hashes.csv"), index=False, encoding="utf-8-sig")

    print(f"[산출] {ws_path}")
    print(f"[산출] {sealed_path}")
    print(f"       시트 {len(wb.sheetnames)}개 · 대조 항목 {len(ITEMS)}개")

    # 누수 검사 — 워크시트에 코드 산출값이 남아 있으면 블라인드가 깨진다
    from openpyxl import load_workbook
    chk = load_workbook(ws_path)
    led = d["ledger"]
    g = led[(led.market == "US") & (led.selection_date == CYCLE)].dropna(subset=["official_adtv90"])
    ktos = float(g[g.security_id == SAMPLE_US].official_adtv90.iloc[0])
    p10 = float(np.percentile(np.sort(g.official_adtv90.values), 10))
    lvl = float(d["idx"].set_index("market_date")["index_level"].loc[pd.Timestamp(INDEX_TEST)])
    banned = {"KTOS ADTV90": ktos, "P10": p10, "지수레벨": lvl,
              "비중 1/6": 1 / 6, "비중 1/18": 1 / 18}
    leaks = []
    for sh in chk.worksheets:
        for row in sh.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)):
                    for k, v in banned.items():
                        if v and abs(c.value - v) < abs(v) * 1e-9:
                            leaks.append(f"{sh.title}!{c.coordinate} = {k}")
    if leaks:
        print("[경고] 코드값 누수 발견 — 블라인드 불성립:")
        for x in leaks:
            print("   ", x)
    else:
        print("[검사] 코드 산출값 누수 없음 — 블라인드 성립")


if __name__ == "__main__":
    main()

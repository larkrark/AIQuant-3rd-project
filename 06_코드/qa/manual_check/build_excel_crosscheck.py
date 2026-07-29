# -*- coding: utf-8 -*-
"""
J-5 엑셀 수식 교차검산본 생성기

무엇인가
  원자료만 싣고, 계산은 **전부 엑셀 수식으로** 걸어 둔 워크북이다.
  파일을 열면 엑셀의 계산 엔진이 원자료에서 최종 지수값까지 직접 계산한다.
  숫자를 하드코딩한 칸은 대조표의 '코드값' 열뿐이다.

왜 만드는가
  같은 규칙을 서로 다른 계산 환경에서 돌려 보면 구현에 기댄 오류가 드러난다.

    파이썬 엔진      engine/           팀 구현
    파이썬 독립본    qa/independent/   규칙 문서 기준 별도 구현
    엑셀 수식        이 파일           제3의 계산 환경
    사람 손          J5 워크시트       가장 독립적

  엑셀은 백분위 보간·부동소수점 처리가 파이썬과 별개로 구현돼 있다.
  네 경로가 같은 값을 내면 산식 해석이 환경에 의존하지 않는다는 뜻이다.

무엇을 증명하지 못하는가
  이 파일의 수식은 내가 룰북을 읽고 쓴 것이라 **규칙 해석은 독립이 아니다.**
  독립 재산출도 같은 해석에서 나왔다. 해석 자체가 틀렸다면 셋 다 같이 틀린다.
  그것을 잡는 것은 사람이 손으로 하는 J-5 검산뿐이다.

주의
  이 파일은 답이 들어 있다. J5_수기검산_워크시트.xlsx 를 다 채운 뒤에 열 것.

실행
  python build_excel_crosscheck.py
"""
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
import paths as P  # noqa: E402

P.force_utf8_stdout()
PILOT = os.path.join(P.DATA, "pilot_run")
PUB = os.path.join(PILOT, "output_f1")
INP = os.path.join(PILOT, "input_krxbm")
OUT = os.path.join(HERE, "J5_엑셀수식_교차검산.xlsx")

CYCLE = "2026-06-30"
CUTOFF = "2026-06-23"
INDEX_BASE = "2026-04-01"
INDEX_TEST = "2026-04-07"
SPLIT_SID, SPLIT_RATIO, SPLIT_EFF = "010120", 5, "2026-04-13"
# SPCX를 마지막에 두어 '시즈닝 통과 8종목'이 연속 범위가 되게 한다
US_ORDER = ["ALAB", "ANET", "APH", "ATI", "ETN", "GEV", "KTOS", "TER", "SPCX"]

NAVY = "1E2761"
C_HEAD = PatternFill("solid", fgColor=NAVY)
C_RAW = PatternFill("solid", fgColor="EFEFE9")
C_CALC = PatternFill("solid", fgColor="E3EEF9")   # 엑셀 수식이 채우는 칸
C_ANS = PatternFill("solid", fgColor="FFE9A8")
C_CODE = PatternFill("solid", fgColor="E6F4EA")
C_WARN = PatternFill("solid", fgColor="FDE8E6")
C_BAND = PatternFill("solid", fgColor="F7F7F4")
F_HEAD = Font(color="FFFFFF", bold=True, size=10)
F_TITLE = Font(bold=True, size=15, color=NAVY)
F_SUB = Font(size=10, color="5A5A55")
F_SEC = Font(bold=True, size=11, color=NAVY)
F_NOTE = Font(size=9, color="6B6B66")
_s = Side("thin", color="C8C7BC")
BOX = Border(left=_s, right=_s, top=_s, bottom=_s)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


def widths(ws, spec):
    for c, w in spec.items():
        ws.column_dimensions[c].width = w


def title(ws, row, text, sub=None):
    ws.cell(row, 1, text).font = F_TITLE
    row += 1
    if sub:
        ws.cell(row, 1, sub).font = F_SUB
        row += 1
    return row + 1


def note(ws, row, lines, ncol=8, fill=C_BAND):
    for ln in lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        c = ws.cell(row, 1, ln)
        c.font = F_NOTE if ln.startswith(("※", "→", "  ")) else Font(size=10)
        c.fill, c.alignment = fill, WRAP
        row += 1
    return row + 1


def head(ws, row, headers, col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row, col + i, h)
        c.fill, c.font, c.border, c.alignment = C_HEAD, F_HEAD, BOX, CTR
    return row + 1


def put(ws, r, c, v, fill=None, fmt=None, align=None, bold=False):
    cell = ws.cell(r, c, v)
    cell.border = BOX
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if bold:
        cell.font = Font(bold=True)
    return cell


def load():
    d = {}
    d["state"] = pd.read_csv(os.path.join(PUB, "daily_market_state.csv"),
                             dtype={"security_id": str}, parse_dates=["market_date"])
    d["cons"] = pd.read_csv(os.path.join(PUB, "constituents_2026-03-31.csv"), dtype={"security_id": str})
    d["w"] = pd.read_csv(os.path.join(PUB, "weights_2026-03-31.csv"), dtype={"security_id": str})
    d["idx"] = pd.read_csv(os.path.join(PUB, "index_vs_benchmark.csv"), parse_dates=["market_date"])
    d["led"] = pd.read_csv(os.path.join(PUB, "adtv90_ledger.csv"), dtype={"security_id": str})
    d["cal"] = pd.read_csv(os.path.join(INP, "calendar.csv"), parse_dates=["market_date"])
    d["fx"] = pd.read_csv(os.path.join(INP, "fx.csv"), parse_dates=["market_date"])
    return d


# ══════════════════════════════════════════════════════════════
def sh_guide(wb):
    ws = wb.create_sheet("00 안내")
    widths(ws, {"A": 20, "B": 26, "C": 26, "D": 24, "E": 20, "F": 18, "G": 16, "H": 14})
    r = title(ws, 1, "J-5 엑셀 수식 교차검산본",
              "계산은 전부 엑셀 수식이다 · 파일을 열면 엑셀이 원자료에서 직접 계산한다")

    r = note(ws, r, [
        "■ 이 파일을 언제 여는가",
        "  J5_수기검산_워크시트.xlsx 를 전부 채우고 저장한 다음입니다. 이 파일에는 답이 들어 있습니다.",
    ], fill=C_WARN)

    r = note(ws, r, [
        "■ 무엇을 하는 파일인가",
        "  원자료만 싣고 계산 칸은 전부 엑셀 수식으로 걸어 두었습니다. 파일을 열면 엑셀이 계산합니다.",
        "  파란 칸(계산)을 클릭하면 수식이 보입니다. 하드코딩된 숫자는 대조표의 '코드값' 열뿐입니다.",
        "",
        "■ 왜 네 경로로 계산하는가",
        "  같은 규칙을 다른 계산 환경에서 돌려 보면 특정 구현에 기댄 오류가 드러납니다.",
    ])

    r = head(ws, r, ["경로", "위치", "무엇이 독립인가", "무엇이 독립이 아닌가"])
    for a, b, c, dd in [
        ("파이썬 엔진", "engine/", "팀의 구현", "—"),
        ("파이썬 독립본", "qa/independent/", "구현이 다름", "규칙 해석이 같음(같은 사람)"),
        ("엑셀 수식", "이 파일", "계산 환경이 다름", "규칙 해석이 같음(같은 사람)"),
        ("사람 손", "J5 워크시트", "해석·계산 모두", "—"),
    ]:
        put(ws, r, 1, a, C_RAW); put(ws, r, 2, b, C_RAW)
        put(ws, r, 3, c, C_RAW); put(ws, r, 4, dd, C_RAW)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 이 파일이 증명하지 못하는 것",
        "  수식은 제가 룰북을 읽고 쓴 것이라 규칙 해석은 독립이 아닙니다. 독립 재산출도 같은 해석에서 나왔습니다.",
        "  해석 자체가 틀렸다면 파이썬 둘과 엑셀이 사이좋게 같이 틀립니다.",
        "  그것을 잡는 것은 사람이 손으로 하는 J-5 검산뿐입니다. 그래서 순서가 중요합니다.",
        "",
        "  엑셀이 잡아 주는 것: 옮겨 적기 오류 · 백분위 보간 구현 차이 · 부동소수점 처리 차이 · 범위 지정 실수",
        "  엑셀이 못 잡는 것:  룰북 오독 · 관측창을 잘못 잡음 · 분모 규칙 오해",
    ], fill=C_WARN)

    r = head(ws, r, ["시트", "내용", "계산 방식"])
    for a, b, c in [
        ("05 계산 자료마감일", "공통 개장일 축 5거래일 역산", "SUMPRODUCT · INDEX/MATCH"),
        ("10 원자료 US거래", "미국 9종목 × 90개장일 종가·거래량·상태", "원자료 (수식 없음)"),
        ("11 계산 ADTV90", "유효관측일수·거래대금합·ADTV90", "COUNTIF · SUMPRODUCT"),
        ("12 계산 P10", "백분위 보간·편입판정", "PERCENTILE"),
        ("20 원자료 가격", "15종목 2일치 종가 + 환율", "원자료 (수식 없음)"),
        ("21 계산 비중", "셀별 종목수 → 1/6 → 종목당", "COUNTIF"),
        ("22 계산 지수", "원화환산 → 가격비 → 가중합", "SUMPRODUCT"),
        ("30 계산 분할", "010120 분할 조정 전후", "산술식"),
        ("99 대조표", "엑셀값 · 수기값 · 코드값", "차이는 수식"),
    ]:
        put(ws, r, 1, a, C_RAW); put(ws, r, 2, b, C_RAW); put(ws, r, 3, c, C_RAW)
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


def sh_cutoff(wb, d):
    """자료마감일 — 공통 개장일 축으로 선정일에서 5거래일 역산 (엑셀 수식)."""
    ws = wb.create_sheet("05 계산 자료마감일")
    widths(ws, {"A": 6, "B": 14, "C": 12, "D": 12, "E": 14, "F": 14, "G": 34})
    r = title(ws, 1, f"자료마감일 — 선정일 {CYCLE} 이전 제5거래일",
              "룰북 §5.2 · D-13 ⑧ · 축은 한·미 공통 개장일")

    r = note(ws, r, [
        "■ 규칙",
        "  선정일 당일은 세지 않습니다(0). 선정일 바로 앞 공통 개장일이 1거래일 전입니다.",
        "  한국과 미국이 모두 열린 날만 셉니다. 한쪽만 열린 날은 번호를 주지 않고 건너뜁니다.",
        "",
        "■ 수식",
        "  공통 개장일  = IF(AND(한국=\"O\", 미국=\"O\"), 1, 0)",
        "  며칠 전      = SUMPRODUCT((날짜범위 > 이 날짜) × (날짜범위 <= 선정일) × 공통플래그범위)",
        "  자료마감일   = INDEX(날짜범위, MATCH(5, 며칠전범위, 0))",
    ], ncol=7)

    cal = d["cal"]
    kr = set(cal[(cal.market == "KR") & (cal.is_market_open == 1)].market_date)
    us = set(cal[(cal.market == "US") & (cal.is_market_open == 1)].market_date)
    days = sorted(x for x in (kr | us)
                  if pd.Timestamp("2026-06-01") <= x <= pd.Timestamp(CYCLE))

    r = head(ws, r, ["", "날짜", "한국 개장", "미국 개장", "공통 개장일", "며칠 전", "메모"])
    f = r
    for i, dt in enumerate(days):
        put(ws, r, 1, i + 1, C_RAW, align=CTR)
        put(ws, r, 2, dt.strftime("%Y-%m-%d"), C_RAW)
        put(ws, r, 3, "O" if dt in kr else "-", C_RAW, align=CTR)
        put(ws, r, 4, "O" if dt in us else "-", C_RAW, align=CTR)
        put(ws, r, 5, f'=IF(AND(C{r}="O",D{r}="O"),1,0)', C_CALC, "#,##0", CTR)
        r += 1
    l = r - 1
    for rr in range(f, l + 1):
        # 자기 날짜보다 뒤이면서 선정일 이하인 공통 개장일 수 = '며칠 전'
        put(ws, rr, 6,
            f'=SUMPRODUCT(($B${f}:$B${l}>B{rr})*($B${f}:$B${l}<="{CYCLE}")*$E${f}:$E${l})',
            C_CALC, "#,##0", CTR)
        put(ws, rr, 7, None, C_RAW)
    r += 1

    put(ws, r, 1, "자료마감일 (며칠 전 = 5 인 날)", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    put(ws, r, 3, f'=INDEX($B${f}:$B${l},MATCH(5,$F${f}:$F${l},0))', C_ANS, "@", CTR)
    ans = r
    r += 1
    put(ws, r, 1, "구간 내 공통 개장일 수", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    put(ws, r, 3, f"=SUM($E${f}:$E${l})", C_CALC, "#,##0", CTR)
    r += 1
    put(ws, r, 1, "한쪽 시장만 열린 날 수", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    put(ws, r, 3, f"=COUNTA($B${f}:$B${l})-SUM($E${f}:$E${l})", C_CALC, "#,##0", CTR)
    put(ws, r, 5, "0이 아니면 축 선택이 결과를 바꾼다", C_RAW)
    r += 2

    r = note(ws, r, [
        "■ 왜 이 시트가 필요한가",
        "  자료마감일이 하루라도 밀리면 ADTV90 관측창이 통째로 이동해 편입 판정까지 달라집니다.",
        "  미래 정보 차단(PIT)의 출발점이라 여기가 틀리면 뒤가 전부 틀립니다.",
        "  실제로 엔진의 F-1 오류가 이 지점이었습니다 — 관측 종료일을 선정일로 잡아 미래 정보가 샜습니다.",
    ], ncol=7)

    ws.sheet_view.showGridLines = False
    return ans


def sh_raw_us(wb, d):
    ws = wb.create_sheet("10 원자료 US거래")
    r = title(ws, 1, "미국 9종목 × 90 개장일 — 원자료",
              f"관측창: 자료마감일 {CUTOFF} 이하 미국 개장일 최근 90일 · 출처 daily_market_state.csv")
    us_open = sorted(d["cal"][(d["cal"].market == "US") & (d["cal"].is_market_open == 1)].market_date)
    win = [x for x in us_open if x <= pd.Timestamp(CUTOFF)][-90:]

    hdr = r
    put(ws, hdr, 1, "날짜", C_HEAD).font = F_HEAD
    ws.column_dimensions["A"].width = 12
    for i, sid in enumerate(US_ORDER):
        c0 = 2 + 3 * i
        ws.merge_cells(start_row=hdr - 1, start_column=c0, end_row=hdr - 1, end_column=c0 + 2)
        m = ws.cell(hdr - 1, c0, sid)
        m.fill, m.font, m.alignment = C_HEAD, F_HEAD, CTR
        for j, nm in enumerate(["종가", "거래량", "상태"]):
            c = ws.cell(hdr, c0 + j, nm)
            c.fill, c.font, c.border, c.alignment = C_HEAD, F_HEAD, BOX, CTR
            ws.column_dimensions[get_column_letter(c0 + j)].width = 13 if j < 2 else 14
    r = hdr + 1
    first = r
    st = d["state"]
    for dt in win:
        put(ws, r, 1, dt.strftime("%Y-%m-%d"), C_RAW)
        for i, sid in enumerate(US_ORDER):
            row = st[(st.security_id == sid) & (st.market_date == dt)]
            c0 = 2 + 3 * i
            if len(row):
                x = row.iloc[0]
                put(ws, r, c0, None if pd.isna(x.raw_close) else float(x.raw_close), C_RAW, "#,##0.00")
                put(ws, r, c0 + 1, None if pd.isna(x.volume) else float(x.volume), C_RAW, "#,##0")
                put(ws, r, c0 + 2, x.daily_market_state, C_RAW, align=CTR)
            else:
                for j in range(3):
                    put(ws, r, c0 + j, None, C_RAW)
        r += 1
    ws.freeze_panes = ws.cell(first, 2)
    ws.sheet_view.showGridLines = False
    return first, r - 1


def sh_calc_adtv(wb, rng):
    a, b = rng
    ws = wb.create_sheet("11 계산 ADTV90")
    widths(ws, {"A": 6, "B": 10, "C": 16, "D": 20, "E": 20, "F": 16, "G": 22, "H": 30})
    r = title(ws, 1, "ADTV90 — 엑셀 수식으로 산출",
              "룰북 §8.1 · 제59조(미국은 원종가×원거래량 재구성) · 정지·무거래는 0 반영, 분모 90")

    r = note(ws, r, [
        "■ 수식",
        "  유효관측일수  = COUNTIF(상태열, \"TRADED\")",
        "  거래대금 합계 = SUMPRODUCT(종가열, 거래량열)     ← 제59조 재구성. 빈 칸은 0으로 처리됨",
        "  ADTV90        = 합계 ÷ 90                        ← 분모는 관측창 길이 90 (결측 없음)",
        "  시즈닝        = IF(유효관측일수 >= 90, \"SEASONED\", \"INCOMPLETE\")",
    ], ncol=8)

    r = head(ws, r, ["", "종목", "유효관측일수", "거래대금 합계", "ADTV90", "시즈닝", "분포 포함값", "메모"])
    first = r
    for i, sid in enumerate(US_ORDER):
        cc, vc, sc = (get_column_letter(2 + 3 * i), get_column_letter(3 + 3 * i),
                      get_column_letter(4 + 3 * i))
        S = "'10 원자료 US거래'!"
        put(ws, r, 1, i + 1, C_RAW, align=CTR)
        put(ws, r, 2, sid, C_RAW, align=CTR)
        put(ws, r, 3, f'=COUNTIF({S}{sc}{a}:{sc}{b},"TRADED")', C_CALC, "#,##0")
        put(ws, r, 4, f'=SUMPRODUCT({S}{cc}{a}:{cc}{b},{S}{vc}{a}:{vc}{b})', C_CALC, "#,##0.00")
        put(ws, r, 5, f"=D{r}/90", C_CALC, "#,##0.00")
        put(ws, r, 6, f'=IF(C{r}>=90,"SEASONED","INCOMPLETE")', C_CALC, align=CTR)
        put(ws, r, 7, f'=IF(F{r}="SEASONED",E{r},"")', C_CALC, "#,##0.00")
        put(ws, r, 8, "시즈닝 미달 — 분포 제외 대상" if sid == "SPCX" else None, C_RAW)
        r += 1
    last = r - 1

    r += 1
    r = note(ws, r, [
        "■ SPCX를 보면 규칙의 의미가 드러납니다",
        "  SPCX는 90일 중 7일만 거래됐습니다. 그런데 합계를 90으로 나눈 값은 오히려 9종목 중 가장 큽니다.",
        "  7일 동안 거래대금이 매우 컸기 때문입니다. 시즈닝 규칙이 없으면 신규 상장주가 유동성 심사를",
        "  가장 여유롭게 통과해 버립니다. 룰북 §8.1이 모집단을 '시즈닝 통과'로 한정한 이유입니다.",
    ], ncol=8)

    ws.sheet_view.showGridLines = False
    return first, last


def sh_calc_p10(wb, rng, d):
    f, l = rng            # 11 시트의 첫/끝 행
    # 시즈닝 판정은 11 시트 G열 수식 =IF(F="SEASONED",E,"") 한 곳에만 둔다.
    # 아래 함수들은 전체 범위를 쓰되, PERCENTILE·SMALL·COUNT·COUNTIF가
    # 텍스트("")를 자동으로 건너뛰므로 결과적으로 시즈닝 통과분만 계산된다.
    # 범위를 손으로 잘라 쓰면 규칙이 두 곳에 흩어져 감사 추적이 어려워진다.
    seasoned_last = l
    ws = wb.create_sheet("12 계산 P10")
    widths(ws, {"A": 34, "B": 22, "C": 18, "D": 40})
    r = title(ws, 1, "P10 하한과 편입 판정 — 엑셀 수식으로 산출",
              "룰북 §8.1(분포·하한) · §9(편입) · PERCENTILE = 선형보간")

    S = "'11 계산 ADTV90'!"
    r = note(ws, r, [
        "■ 엑셀의 PERCENTILE 은 파이썬 numpy.percentile 기본값과 같은 선형보간을 씁니다.",
        "  위치 h = (n − 1) × p, 소수부는 아래위 두 값 사이를 비례 배분합니다.",
        "  두 환경이 같은 값을 내면 보간 구현이 환경에 의존하지 않는다는 뜻입니다.",
    ], ncol=4)

    rows = [
        ("분포 모집단 n — 시즈닝 통과만", f"=COUNT({S}G{f}:G{l})", "#,##0", "룰북 §8.1 '보통주·시즈닝 통과'"),
        ("보간 위치 h = (n−1) × 0.10", "=(B{0}-1)*0.1", "0.00", "h < 1 이면 P10은 최솟값과 2번째 사이"),
        ("P10 하한 (시즈닝 통과 8종목)", f"=PERCENTILE({S}G{f}:G{seasoned_last},0.1)", "#,##0.00", "★ 공식 산출값"),
        ("최솟값 x1", f"=SMALL({S}G{f}:G{seasoned_last},1)", "#,##0.00", ""),
        ("두 번째 값 x2", f"=SMALL({S}G{f}:G{seasoned_last},2)", "#,##0.00", ""),
        ("보간 검산 = x1 + (h−1의내림)×(x2−x1)", "=B{3}+(B{1}-INT(B{1}))*(B{4}-B{3})", "#,##0.00",
         "PERCENTILE 과 같아야 함"),
        ("KTOS ADTV90", f"=INDEX({S}E{f}:E{l},MATCH(\"KTOS\",{S}B{f}:B{l},0))", "#,##0.00", ""),
        ("KTOS 편입 여부", "=IF(B{6}>=B{2},\"편입\",\"제외\")", "@", "P10 이상이면 편입"),
        ("KTOS의 P10 대비 차이 (%)", "=(B{6}/B{2}-1)*100", "0.00", ""),
        ("P10 미만 종목 수", f"=COUNTIF({S}G{f}:G{seasoned_last},\"<\"&B{{2}})", "#,##0", "몇 종목이 떨어지는가"),
        ("2종목 이상 탈락에 필요한 최소 n", "=CEILING(1/0.1,1)+1", "#,##0", "h=(n−1)×0.1 ≥ 1"),
    ]
    start = r
    r = head(ws, r, ["항목", "값 (엑셀 수식)", "", "근거·비고"])
    base = r
    for i, (name, formula, fmt, memo) in enumerate(rows):
        put(ws, r, 1, name, C_RAW)
        fx = formula.format(*[base + k for k in range(len(rows))])
        put(ws, r, 2, fx, C_ANS if "★" in memo else C_CALC, fmt)
        put(ws, r, 4, memo, C_RAW)
        r += 1
    r += 1

    r = note(ws, r, [
        "■ 참고 — SPCX를 모집단에 넣으면 어떻게 되는가 (규칙상 넣지 않음)",
    ], ncol=4)
    put(ws, r, 1, "P10 (SPCX 포함 9종목)", C_RAW)
    put(ws, r, 2, f"=PERCENTILE({S}E{f}:E{l},0.1)", C_CALC, "#,##0.00")
    put(ws, r, 4, "n이 9가 되어 h가 0.8로 커진다. 하위 두 값은 그대로라 P10만 올라간다.", C_RAW)
    r += 2

    r = note(ws, r, [
        "■ 여기서 나오는 구조 문제",
        "  h가 0과 1 사이면 P10은 항상 최솟값과 두 번째 값 사이에 놓입니다.",
        "  따라서 P10 미만 종목은 언제나 정확히 한 개입니다. 그 종목의 거래대금 크기와 무관합니다.",
        "  → 지금 P10은 '유동성이 일정 수준 미만이면 제외'가 아니라 '가장 낮은 하나를 제외'로 작동합니다.",
    ], ncol=4)

    ws.sheet_view.showGridLines = False
    return base


def sh_raw_price(wb, d):
    ws = wb.create_sheet("20 원자료 가격")
    widths(ws, {"A": 6, "B": 12, "C": 8, "D": 18, "E": 20, "F": 16, "G": 16})
    r = title(ws, 1, "구성종목 종가 2일치 + 환율 — 원자료",
              f"기준일 {INDEX_BASE} · 검산일 {INDEX_TEST} · 출처 daily_market_state.csv · fx.csv")

    fxs = d["fx"].set_index("market_date")["fx_rate"]
    b, t = pd.Timestamp(INDEX_BASE), pd.Timestamp(INDEX_TEST)
    r = head(ws, r, ["구분", "날짜", "환율(원/달러)"])
    fx_rows = {}
    for lab, dt in [("기준일", b), ("검산일", t)]:
        put(ws, r, 1, lab, C_RAW)
        put(ws, r, 2, dt.strftime("%Y-%m-%d"), C_RAW)
        put(ws, r, 3, float(fxs.loc[dt]), C_RAW, "#,##0.0")
        fx_rows[lab] = r
        r += 1
    r += 1

    r = head(ws, r, ["", "종목", "지역", "테마", "셀", f"기준일 종가", f"검산일 종가"])
    first = r
    st = d["state"]
    cons = d["cons"]
    sel = cons[cons.selected_flag == 1].sort_values(["market", "security_id"])
    for i, (_, row) in enumerate(sel.iterrows()):
        px = st[st.security_id == row.security_id].set_index("market_date")["raw_close"]
        put(ws, r, 1, i + 1, C_RAW, align=CTR)
        put(ws, r, 2, row.security_id, C_RAW, align=CTR)
        put(ws, r, 3, row.market, C_RAW, align=CTR)
        put(ws, r, 4, row.primary_theme, C_RAW)
        put(ws, r, 5, row.cell_id, C_RAW)
        put(ws, r, 6, float(px.loc[b]), C_RAW, "#,##0.00")
        put(ws, r, 7, float(px.loc[t]), C_RAW, "#,##0.00")
        r += 1
    ws.sheet_view.showGridLines = False
    return first, r - 1, fx_rows


def sh_calc_weight(wb, prng):
    f, l, _ = prng
    ws = wb.create_sheet("21 계산 비중")
    widths(ws, {"A": 6, "B": 22, "C": 16, "D": 16, "E": 18, "F": 30})
    r = title(ws, 1, "6셀 목표비중 — 엑셀 수식으로 산출",
              "D-04 ②(테마 1:1:1) · D-10 ①(지역 50:50) · D-10 ②(6셀 각 1/6)")

    R = "'20 원자료 가격'!"
    r = note(ws, r, [
        "■ 수식",
        "  셀 목표비중 = 1/6                                    ← D-10 ② 명문",
        "  편입 종목수 = COUNTIF(셀 열, 해당 셀)",
        "  종목당 비중 = 셀 목표비중 ÷ 편입 종목수",
        "",
        "  ※ 6셀 각 1/6은 테마 1:1:1과 지역 50:50을 만족하는 유일한 해가 아닙니다.",
        "    미지수 6개에 독립 제약이 4개라 자유도가 2 남습니다. 근거는 유일성이 아니라 D-10 ② 명문입니다.",
    ], ncol=6)

    r = head(ws, r, ["", "셀", "셀 목표비중", "편입 종목수", "종목당 비중", "메모"])
    cf = r
    cells = ["KR_AI_ROBOTICS", "KR_ENERGY_POWER", "KR_SPACE_DEFENSE",
             "US_AI_ROBOTICS", "US_ENERGY_POWER", "US_SPACE_DEFENSE"]
    for i, cell in enumerate(cells):
        put(ws, r, 1, i + 1, C_RAW, align=CTR)
        put(ws, r, 2, cell, C_RAW)
        put(ws, r, 3, "=1/6", C_CALC, "0.000000")
        put(ws, r, 4, f'=COUNTIF({R}$E${f}:$E${l},B{r})', C_CALC, "#,##0")
        put(ws, r, 5, f"=C{r}/D{r}", C_CALC, "0.000000")
        put(ws, r, 6, None, C_RAW)
        r += 1
    cl = r - 1
    put(ws, r, 2, "합계", C_RAW, bold=True)
    put(ws, r, 3, f"=SUM(C{cf}:C{cl})", C_ANS, "0.000000")
    put(ws, r, 4, f"=SUM(D{cf}:D{cl})", C_CALC, "#,##0")
    put(ws, r, 6, "정확히 1.000000 이어야 함", C_RAW)
    tot = r
    r += 2

    r = head(ws, r, ["", "종목", "셀", "목표비중", "", "메모"])
    sf = r
    for i in range(f, l + 1):
        put(ws, r, 1, i - f + 1, C_RAW, align=CTR)
        put(ws, r, 2, f"={R}B{i}", C_CALC, align=CTR)
        put(ws, r, 3, f"={R}E{i}", C_CALC)
        put(ws, r, 4, f"=VLOOKUP(C{r},$B${cf}:$E${cl},4,FALSE)", C_CALC, "0.000000")
        put(ws, r, 6, None, C_RAW)
        r += 1
    sl = r - 1
    put(ws, r, 3, "합계", C_RAW, bold=True)
    put(ws, r, 4, f"=SUM(D{sf}:D{sl})", C_ANS, "0.000000")
    r += 2

    put(ws, r, 1, "최대 비중", C_RAW)
    put(ws, r, 2, f"=MAX(D{sf}:D{sl})", C_ANS, "0.000000")
    put(ws, r, 3, f'=INDEX(B{sf}:B{sl},MATCH(MAX(D{sf}:D{sl}),D{sf}:D{sl},0))', C_CALC, align=CTR)
    put(ws, r, 4, f"=B{r}*100", C_CALC, "0.0000")
    put(ws, r, 6, "% 표기", C_RAW)
    maxrow = r

    ws.sheet_view.showGridLines = False
    return sf, sl, tot, maxrow


def sh_calc_index(wb, prng, wrng):
    pf, pl, fx_rows = prng
    sf, sl, _, _ = wrng
    ws = wb.create_sheet("22 계산 지수")
    widths(ws, {"A": 6, "B": 12, "C": 8, "D": 16, "E": 18, "F": 18, "G": 16, "H": 18, "I": 16})
    r = title(ws, 1, f"지수 레벨 {INDEX_TEST} — 엑셀 수식으로 산출",
              "룰북 §13 · D-08 · 기준값 1,000 · 원화 무헤지 · SAME_DAY_ECOS")

    R = "'20 원자료 가격'!"
    W = "'21 계산 비중'!"
    r = note(ws, r, [
        "■ 수식",
        "  원화 가격  = IF(지역=\"US\", 종가 × 그 날짜의 환율, 종가)     ← 한국 종목엔 곱하지 않음",
        "  가격비     = 검산일 원화가격 ÷ 기준일 원화가격",
        "  지수       = SUMPRODUCT(목표비중, 가격비) × 1000",
        "",
        "  ※ 기준일과 검산일에 서로 다른 환율을 씁니다. 같은 환율을 쓰면 환율 변동분이 사라집니다.",
    ], ncol=9)

    r = head(ws, r, ["", "종목", "지역", "목표비중", "기준일 원화가격", "검산일 원화가격",
                     "가격비", "비중×가격비", "메모"])
    f = r
    for i in range(pf, pl + 1):
        k = i - pf
        put(ws, r, 1, k + 1, C_RAW, align=CTR)
        put(ws, r, 2, f"={R}B{i}", C_CALC, align=CTR)
        put(ws, r, 3, f"={R}C{i}", C_CALC, align=CTR)
        put(ws, r, 4, f"={W}D{sf + k}", C_CALC, "0.000000")
        put(ws, r, 5, f'=IF(C{r}="US",{R}F{i}*{R}$C${fx_rows["기준일"]},{R}F{i})', C_CALC, "#,##0.00")
        put(ws, r, 6, f'=IF(C{r}="US",{R}G{i}*{R}$C${fx_rows["검산일"]},{R}G{i})', C_CALC, "#,##0.00")
        put(ws, r, 7, f"=F{r}/E{r}", C_CALC, "0.00000000")
        put(ws, r, 8, f"=D{r}*G{r}", C_CALC, "0.00000000")
        put(ws, r, 9, None, C_RAW)
        r += 1
    l = r - 1
    put(ws, r, 7, "합계", C_RAW, bold=True)
    put(ws, r, 8, f"=SUM(H{f}:H{l})", C_CALC, "0.00000000")
    tot = r
    r += 2

    put(ws, r, 1, f"{INDEX_TEST} 지수 레벨 = 합계 × 1000", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    put(ws, r, 4, f"=H{tot}*1000", C_ANS, "#,##0.0000")
    idx_row = r

    ws.sheet_view.showGridLines = False
    return idx_row


def sh_calc_split(wb, d, wrng):
    sf, sl, _, _ = wrng
    ws = wb.create_sheet("30 계산 분할")
    widths(ws, {"A": 6, "B": 14, "C": 14, "D": 18, "E": 18, "F": 16, "G": 34})
    r = title(ws, 1, f"{SPLIT_SID} 액면분할 5:1 — 조정 전후 비교",
              f"데이터사전 4.1 · 경계일 {SPLIT_EFF}(신주권상장일) · DART 20260205800571")

    r = note(ws, r, [
        "■ 수식",
        f"  조정 종가 = IF(날짜 < {SPLIT_EFF}, 원종가 ÷ 5, 원종가)",
        "  수익률    = 당일 조정종가 ÷ 전일 조정종가 − 1",
        "",
        "  ※ 거래정지 기간(04-08~10)의 종가도 분할 전 기준이므로 함께 나눕니다.",
        f"  ※ 공시에는 효력발생일 2026-04-10 과 상장일 {SPLIT_EFF} 이 따로 적혀 있습니다.",
        "    가격 계열에서는 거래가 실제로 재개된 상장일이 경계입니다.",
    ], ncol=7)

    st = d["state"]
    sub = st[(st.security_id == SPLIT_SID) &
             (st.market_date >= "2026-04-06") & (st.market_date <= "2026-04-16")].sort_values("market_date")
    r = head(ws, r, ["", "날짜", "원종가", "상태", "조정 종가", "조정 후 수익률", "조정 전 수익률"])
    f = r
    for i, (_, row) in enumerate(sub.iterrows()):
        dt = row.market_date.strftime("%Y-%m-%d")
        put(ws, r, 1, i + 1, C_RAW, align=CTR)
        put(ws, r, 2, dt, C_RAW)
        put(ws, r, 3, float(row.raw_close), C_RAW, "#,##0")
        put(ws, r, 4, row.daily_market_state, C_RAW, align=CTR)
        put(ws, r, 5, f'=IF(B{r}<"{SPLIT_EFF}",C{r}/{SPLIT_RATIO},C{r})', C_CALC, "#,##0.0")
        if i:
            put(ws, r, 6, f"=E{r}/E{r-1}-1", C_CALC, "0.00%")
            put(ws, r, 7, f"=C{r}/C{r-1}-1", C_CALC, "0.00%")
        else:
            put(ws, r, 6, None, C_RAW)
            put(ws, r, 7, None, C_RAW)
        r += 1
    r += 1

    eff_row = f + list(sub.market_date.dt.strftime("%Y-%m-%d")).index(SPLIT_EFF)
    put(ws, r, 1, "① 조정 안 했을 때 04-13 수익률", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    put(ws, r, 4, f"=G{eff_row}", C_ANS, "0.00%")
    put(ws, r, 6, "실재하지 않는 하락", C_RAW)
    r += 1
    put(ws, r, 1, "② 조정 했을 때 04-13 수익률", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    put(ws, r, 4, f"=F{eff_row}", C_ANS, "0.00%")
    put(ws, r, 6, "0%가 나오면 과보정", C_WARN)
    r += 1
    put(ws, r, 1, "③ 차이 (① − ②)", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    put(ws, r, 4, f"=D{r-2}-D{r-1}", C_CALC, "0.00%")
    r += 1
    put(ws, r, 1, "④ 04-13 '당일' 지수 기여 차이 = ③ × 010120 목표비중", C_RAW)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    put(ws, r, 4,
        f"=D{r-1}*VLOOKUP(\"{SPLIT_SID}\",'21 계산 비중'!$B${sf}:$D${sl},3,FALSE)", C_CALC, "0.00%")
    put(ws, r, 6, "★ 전 구간 효과가 아님 — 아래 주의 참조", C_WARN)
    r += 2

    r = note(ws, r, [
        "■ 자기 점검",
        "  조정 후 04-13 수익률이 0%면 두 번 나눈 것입니다. 올바르면 플러스가 나옵니다.",
        "  분할 기준가보다 높게 재개됐다는 뜻이고, 그것이 실제 시장 반응입니다.",
    ], ncol=7, fill=C_WARN)

    r = note(ws, r, [
        "■ 주의 — ④를 '지수가 6.66%p 틀렸다'의 근거로 쓰면 안 됩니다. 축이 다릅니다.",
        "",
        "  ④는 04-13 '하루'의 기여 차이입니다. 분할은 하루 사건이지만, 기준일 고정가중 구조에서는",
        "  조정이 그 이후 전 구간의 가격비를 5배로 바꿉니다. 효과가 하루에 갇히지 않습니다.",
        "",
        "  올바른 전 구간 분해는 이렇습니다.",
        "    010120 전 구간(04-01→06-30) 수익률   조정 전 −70.03% / 조정 후 +49.87% → 차이 119.90%p",
        "    119.90%p × 목표비중 5.5556%(1/18) = 6.66%p = 지수 누적수익률 차이",
        "",
        "  그리고 상대 수익률 차(1337.75/1271.14 − 1 = 5.24%)는 또 다른 축입니다. 셋을 섞지 마십시오.",
        "  이 워크북은 04-07 지수만 검산하므로 전 구간 효과는 대상이 아닙니다.",
        "  전 구간 수치는 qa/corrected_run/ 산출물에 있습니다.",
    ], ncol=7)

    ws.sheet_view.showGridLines = False
    return f, eff_row


def sh_compare(wb, d, refs):
    ws = wb.create_sheet("99 대조표")
    widths(ws, {"A": 6, "B": 34, "C": 22, "D": 22, "E": 22, "F": 16, "G": 16, "H": 34})
    r = title(ws, 1, "세 경로 대조 — 엑셀 수식 · 수기 · 코드",
              "엑셀값은 이 파일의 수식이 계산한 값 · 수기값은 직접 옮겨 적으십시오")

    r = note(ws, r, [
        "■ 보는 법",
        "  엑셀값과 코드값이 같으면 산식 해석이 계산 환경에 의존하지 않는다는 뜻입니다.",
        "  수기값이 둘과 다르면, 사람이 틀렸을 수도 있고 규칙 해석이 갈렸을 수도 있습니다. 원인을 적으십시오.",
        "  세 값이 모두 같아도 규칙 자체를 잘못 읽었을 가능성은 남습니다 — 그건 룰북 조항 대조로만 확인됩니다.",
    ])

    led = d["led"]
    g = led[(led.market == "US") & (led.selection_date == "2026-06-30")]
    ktos = float(g[g.security_id == "KTOS"].official_adtv90.iloc[0])
    import numpy as np
    v = np.sort(g.dropna(subset=["official_adtv90"]).official_adtv90.values)
    p10 = float(np.percentile(v, 10))
    idx = float(d["idx"].set_index("market_date")["index_level"].loc[pd.Timestamp(INDEX_TEST)])
    wmax = float(d["w"].final_target_weight.max())
    st = d["state"]
    px = st[st.security_id == SPLIT_SID].set_index("market_date")["raw_close"]
    p0, p1 = float(px.loc[pd.Timestamp("2026-04-07")]), float(px.loc[pd.Timestamp(SPLIT_EFF)])

    P10R, IDXR, WR, SPR = refs["p10"], refs["idx"], refs["w"], refs["split"]
    items = [
        ("A", "자료마감일", f"='05 계산 자료마감일'!C{refs['cut']}", CUTOFF, "@"),
        ("B", "KTOS ADTV90 (USD)", f"='12 계산 P10'!B{P10R+6}", ktos, "#,##0.00"),
        ("C", "ADTV90 분모 (개장일수)", "=90", 90, "#,##0"),
        ("D", "미국 P10 하한 (USD)", f"='12 계산 P10'!B{P10R+2}", p10, "#,##0.00"),
        ("E", "KTOS 편입 여부", f"='12 계산 P10'!B{P10R+7}", "제외", "@"),
        ("F", "P10 미만 종목 수", f"='12 계산 P10'!B{P10R+9}", 1, "#,##0"),
        ("G", "2종목 이상 탈락 최소 n", f"='12 계산 P10'!B{P10R+10}", 11, "#,##0"),
        ("H", "목표비중 합계", f"='21 계산 비중'!C{WR}", 1.0, "0.000000"),
        ("I", "최대 비중 (%)", f"='21 계산 비중'!D{refs['wmax']}", wmax * 100, "0.0000"),
        ("J", f"{INDEX_TEST} 지수 레벨", f"='22 계산 지수'!D{IDXR}", idx, "#,##0.0000"),
        ("K", "04-13 조정 전 수익률", f"='30 계산 분할'!D{SPR}", p1 / p0 - 1, "0.00%"),
        ("L", "04-13 조정 후 수익률", f"='30 계산 분할'!D{SPR+1}", p1 / (p0 / SPLIT_RATIO) - 1, "0.00%"),
    ]
    r = head(ws, r, ["", "검산 항목", "엑셀 수식값", "수기값 (직접 기입)", "코드값", "엑셀−코드", "판정", "메모"])
    first = r
    for code, name, formula, codeval, fmt in items:
        put(ws, r, 1, code, C_RAW, align=CTR)
        put(ws, r, 2, name, C_RAW)
        put(ws, r, 3, formula, C_CALC, fmt)
        put(ws, r, 4, None, PatternFill("solid", fgColor="FFF7D1"), fmt)
        put(ws, r, 5, codeval, C_CODE, fmt)
        if isinstance(codeval, (int, float)):
            put(ws, r, 6, f"=C{r}-E{r}", C_CALC, "0.00E+00")
            put(ws, r, 7, f'=IF(ABS(C{r}-E{r})<0.000001,"일치","★차이★")', C_CALC, align=CTR)
        else:
            put(ws, r, 6, "—", C_RAW, align=CTR)
            put(ws, r, 7, f'=IF(C{r}=E{r},"일치","★차이★")', C_CALC, align=CTR)
        put(ws, r, 8, None, C_RAW)
        r += 1
    last = r - 1
    r += 1
    put(ws, r, 2, "엑셀 = 코드 일치 항목 수", C_RAW, bold=True)
    put(ws, r, 3, f'=COUNTIF(G{first}:G{last},"일치")', C_ANS, "#,##0")
    put(ws, r, 5, f"/ {len(items)}", C_RAW)

    ws.freeze_panes = ws.cell(first, 1)
    ws.sheet_view.showGridLines = False


def main():
    d = load()
    wb = Workbook()
    wb.remove(wb.active)

    sh_guide(wb)
    cut_row = sh_cutoff(wb, d)
    us_rng = sh_raw_us(wb, d)
    adtv_rng = sh_calc_adtv(wb, us_rng)
    p10_base = sh_calc_p10(wb, adtv_rng, d)
    prng = sh_raw_price(wb, d)
    wrng = sh_calc_weight(wb, prng)
    idx_row = sh_calc_index(wb, prng, wrng)
    sf_row, eff_row = sh_calc_split(wb, d, wrng)
    split_ans = eff_row + (len(range(0, 0)) or 0)
    # 30 시트 답 칸: 표 끝 다음 블록. sh_calc_split 에서 ①이 놓인 행을 계산
    split_first_ans = None
    ws30 = wb["30 계산 분할"]
    for rr in range(1, ws30.max_row + 1):
        v = ws30.cell(rr, 1).value
        if isinstance(v, str) and v.startswith("① 조정"):
            split_first_ans = rr
            break

    refs = {"cut": cut_row, "p10": p10_base, "idx": idx_row, "w": wrng[2],
            "wmax": wrng[3], "split": split_first_ans}
    sh_compare(wb, d, refs)

    wb.save(OUT)
    print(f"[산출] {OUT}")
    print(f"       시트 {len(wb.sheetnames)}개 — {', '.join(wb.sheetnames)}")
    print("       계산 칸은 전부 엑셀 수식. 파일을 열면 엑셀이 계산합니다.")


if __name__ == "__main__":
    main()

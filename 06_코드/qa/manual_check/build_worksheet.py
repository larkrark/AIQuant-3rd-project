# -*- coding: utf-8 -*-
"""수기 검산(J-5) 워크시트 생성 — 원자료를 채우고 계산 칸은 비워 둔다.

수기 검산의 목적은 사람이 원자료에서 최종값까지 손으로 따라가는 것이다(룰북 §14 "수기·코드
교차검산" · §18.4). 따라서 이 스크립트는 **숫자를 옮겨 적는 일만** 한다 — 계산·판정 칸은
비운 채로 두고, 사람이 엑셀 수식으로 직접 채운다. 여기서 계산까지 해버리면 세 번째 코드
구현일 뿐이고 검산이 되지 않는다.

산출: manual_check/J5_수기검산_워크시트.xlsx   ← 사람이 채우는 본 (코드 산출값 없음)
      manual_check/J5_코드값_봉인.xlsx        ← 대조용 정답지 (수기 채움 완료 후에 열 것)
      manual_check/snapshot_hashes.csv

★ 블라인드 구조: 코드 산출값을 워크시트에 나란히 두면 답을 보며 채우게 되어 검산이 성립하지
   않는다(확증편향). 그래서 코드값은 **별도 파일**로 분리한다. 수기 칸을 모두 채우고 저장한
   뒤에 봉인 파일을 열어 5_대조표에 옮겨 적는다.
   ※ 이 분리는 "블라인드 수기 검산"을 성립시키는 조건이며, 독립 재산출(qa/independent)은
      산출물 사전 열람이 있었으므로 블라인드가 아니다 — 두 검증의 성격이 다르다.

표본 근거(룰북 §18.2 "경계사례를 포함한 표본"):
  주 표본 010120 × 2026-04-07~04-14 — 거래정지 3일(0 반영 vs 유효관측일 제외가 갈리는 지점)
                                      + 04-13 분할(수정주가 부재 문제가 실제로 터지는 날)
  보조   KTOS(하한 대비 -0.5% 경계) · SPCX(시즈닝 미달) · 2026-03-31 리밸 전체 18종목
"""
import os
import sys
import json
import hashlib
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import paths as P

P.force_utf8_stdout()
HERE = os.path.dirname(os.path.abspath(__file__))

SAMPLE_SID = "010120"
SAMPLE_DAYS = ["2026-04-07", "2026-04-08", "2026-04-09", "2026-04-10", "2026-04-13", "2026-04-14"]
REBAL = "2026-03-31"
INDEX_DAYS = ["2026-04-07", "2026-04-08", "2026-04-09", "2026-04-10", "2026-04-13", "2026-04-14"]

# --- 서식 ---
H = PatternFill("solid", fgColor="1E2761")          # 머리행 (팀 네이비)
FILL_GIVEN = PatternFill("solid", fgColor="EEEEE8")  # 주어진 원자료 (건드리지 말 것)
FILL_INPUT = PatternFill("solid", fgColor="FFF9DB")  # 사람이 채울 칸
FILL_CODE = PatternFill("solid", fgColor="E8F0FA")   # 코드 산출값 (대조 상대)
FILL_NOTE = PatternFill("solid", fgColor="FDECEA")   # 주의 칸
F_H = Font(color="FFFFFF", bold=True, size=10)
F_T = Font(bold=True, size=12, color="1E2761")
THIN = Border(*[Side("thin", color="C3C2B7")] * 4)


def _title(ws, row, text, sub=""):
    ws.cell(row, 1, text).font = F_T
    if sub:
        ws.cell(row + 1, 1, sub).font = Font(size=9, color="52514E")
    return row + (3 if sub else 2)


def _header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c)
        cell.fill, cell.font, cell.border = H, F_H, THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(widths or [], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1


def _row(ws, row, values, fills):
    for i, (v, f) in enumerate(zip(values, fills), 1):
        cell = ws.cell(row, i, v)
        cell.border = THIN
        if f:
            cell.fill = f
    return row + 1


# ─────────────────────────────────────────────────────────────
def sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def snapshot(input_dir: str, out_csv: str) -> pd.DataFrame:
    """2단계 — 원자료 스냅샷 고정. 손계산 시작 전 입력을 동결한다(담당표 '원자료 스냅샷 보존')."""
    rows = []
    for n in sorted(os.listdir(input_dir)):
        p = os.path.join(input_dir, n)
        if os.path.isfile(p):
            rows.append({"file": n, "bytes": os.path.getsize(p), "sha256": sha256(p)})
    df = pd.DataFrame(rows)
    df.to_csv(out_csv, index=False)
    return df


# ─────────────────────────────────────────────────────────────
def sheet_cover(wb, snap, commit):
    ws = wb.create_sheet("0_표지")
    r = _title(ws, 1, "J-5 수기 검산 워크시트",
               "룰북 §14 산출 정합성(수기·코드 교차검산) · §18.4 승인기준 / 담당 김민호(QA)")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 78

    for k, v in [("대상 rule_version", "v0.9-pilot"),
                 ("입력 스냅샷", "06_코드/data/pilot_run/input_krxbm/"),
                 ("커밋 SHA", commit),
                 ("주 표본", f"{SAMPLE_SID} × {SAMPLE_DAYS[0]}~{SAMPLE_DAYS[-1]}"),
                 ("표본 선정 근거", "룰북 §18.2 경계사례 포함 — 거래정지 3일 + 04-13 분할"),
                 ("보조 표본", "KTOS(하한 -0.5% 경계) · SPCX(시즈닝 미달) · 2026-03-31 리밸 18종목"),
                 ("검산 도구", "엑셀 수식만. 파이썬·스크립트 사용 금지(사용 시 세 번째 코드 구현일 뿐)"),
                 ("표시 자릿수", "금액 소수 4자리 · 비중 소수 9자리 · 지수 소수 6자리"),
                 ("작성일", ""), ("작성자", "")]:
        ws.cell(r, 1, k).font = Font(bold=True, size=10)
        c = ws.cell(r, 2, v)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if v == "":
            c.fill = FILL_INPUT
        r += 1

    r += 1
    ws.cell(r, 1, "허용오차 (문서 규정 없음 — 제 제안, 팀 승인 필요)").font = F_T
    r = _header(ws, r + 1, ["항목", "허용오차", "근거"], [26, 20, 78])
    for a, b, c in [("상태코드·편입판정", "완전일치", "범주값 — 오차 개념 없음"),
                    ("거래대금·ADTV90", "상대 1e-6", "사람 반올림 흡수"),
                    ("목표비중", "절대 1e-9 · 총합 1.0", "룰북 §14 비중합계 100%"),
                    ("지수레벨", "상대 1e-6", "누적 곱이라 상대기준")]:
        r = _row(ws, r, [a, b, c], [None, FILL_INPUT, None])

    r += 1
    ws.cell(r, 1, "입력 스냅샷 해시 (SHA-256)").font = F_T
    r = _header(ws, r + 1, ["파일", "바이트", "sha256"], [26, 14, 78])
    for x in snap.itertuples():
        r = _row(ws, r, [x.file, x.bytes, x.sha256], [FILL_GIVEN] * 3)

    r += 1
    ws.cell(r, 1, "칸 색 규칙").font = F_T
    for label, fill in [("주어진 원자료 — 수정 금지", FILL_GIVEN),
                        ("사람이 채울 칸", FILL_INPUT),
                        ("코드 산출값 — 대조 상대", FILL_CODE),
                        ("주의 — 근거 조항이 없을 수 있는 칸", FILL_NOTE)]:
        r += 1
        ws.cell(r, 1).fill = fill
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2, label)


def sheet_l1(wb, prices, halts, cal):
    """L1 — 일별 상태코드·거래대금. 미결 없음."""
    ws = wb.create_sheet("1_L1_상태코드")
    r = _title(ws, 1, f"L1 · 일별 상태코드와 거래대금 — {SAMPLE_SID}",
               "룰북 §8.1(상태코드 6종·제59조 복원) · R6(NA≠0) / 미결 없음 — 조항만으로 판정 가능")
    r = _header(ws, r, ["일자", "개장?", "원종가", "거래량", "KRX 거래대금", "정지등록",
                        "① 상태코드", "② 근거 조항", "③ 일별 거래대금", "④ 산출 경로",
                        "⑤ 유효관측일 포함?"],
                [12, 8, 12, 14, 18, 10, 16, 26, 18, 18, 16])
    kr_open = set(cal[(cal.market == "KR") & (cal.is_market_open == 1)].market_date.astype(str))
    hs = {str(x.market_date) for x in halts.itertuples()
          if x.security_id == SAMPLE_SID and str(x.full_day_halt).strip().lower() in ("true", "1", "y")}
    for d in SAMPLE_DAYS:
        g = prices[(prices.security_id == SAMPLE_SID) & (prices.market_date == d)]
        px = float(g.raw_close.iloc[0]) if len(g) else None
        vol = float(g.volume.iloc[0]) if len(g) else None
        etv = float(g.exchange_trading_value.iloc[0]) if len(g) and pd.notna(
            g.exchange_trading_value.iloc[0]) else None
        r = _row(ws, r, [d, "O" if d in kr_open else "X", px, vol, etv,
                         "O" if d in hs else "", None, None, None, None, None],
                 [FILL_GIVEN] * 6 + [FILL_INPUT] * 5)
    r += 1
    for t in ["※ ①은 룰북 §8.1 6종 중 하나. TRADING_HALT 가 ZERO_VOLUME 보다 우선한다(§8.1 명시).",
              "※ ②에 조항 번호를 반드시 적는다. 적을 조항이 없으면 그 칸이 미결 후보다.",
              "※ ③ 정지일은 0 반영(D-13 ①), 자료 결측은 0이 아니라 NA(R6). 둘을 섞지 않는다.",
              "※ ⑤ 시즈닝 유효관측일 = 개장 + 상장 중 + 정지 아님. 무거래일은 포함, 정지일은 제외."]:
        ws.cell(r, 1, t).font = Font(size=9, color="52514E")
        r += 1


L2_ITEMS = [("관측창 개장일수 (목표 90)", "observed_open_days"),
            ("그중 정지일수", "halt_days_90"),
            ("그중 무거래일수", "zero_volume_days_90"),
            ("그중 자료결측일수", "missing_days_90"),
            ("ADTV90 분모 (90 − 결측일수)", None),
            ("ADTV90 분자 (거래대금 합)", None),
            ("공식 ADTV90 (정지일 0 반영)", "official_adtv90"),
            ("진단값 (정지일 분모 제외)", "adtv90_exclude_halt"),
            ("시즈닝 유효관측일수", "seasoning_days"),
            ("시즈닝 판정 (≥90)", "seasoning_status"),
            ("KR 시장 P10 하한", None)]


def sheet_l2(wb):
    """L2 — ADTV90·시즈닝·P10. 코드값은 봉인 파일에 있다(블라인드)."""
    ws = wb.create_sheet("2_L2_ADTV90")
    r = _title(ws, 1, f"L2 · ADTV90 · 시즈닝 · P10 하한 — {SAMPLE_SID}",
               "룰북 §8.1 확정 계산명세 / 가정 U-1(상태코드 우선순위)·U-2(백분위 보간) 명시 필요")
    r = _header(ws, r, ["항목", "① 수기 계산값", "② 근거 조항"], [34, 22, 34])
    for label, _ in L2_ITEMS:
        r = _row(ws, r, [label, None, None], [None, FILL_INPUT, FILL_INPUT])
    r += 1
    for t in ["※ 분모 = 90 − 자료결측일수 (룰북 R6). 정지일은 분모에 남기고 값만 0으로 넣는다.",
              "※ P10 은 시장별 시즈닝 통과 종목 모집단으로 산출(룰북 §8.1). 보간 방식은 문서 미규정 = U-2.",
              "※ 90일 원자료는 '2b_L2_원자료' 시트에 있다. SUM·COUNTIF 로 직접 집계할 것.",
              "※ 코드 산출값은 J5_코드값_봉인.xlsx 에 있다. 이 시트를 다 채운 뒤에 열 것."]:
        ws.cell(r, 1, t).font = Font(size=9, color="52514E")
        r += 1


def sheet_l2_raw(wb, states_code):
    """L2 집계용 원자료 — 관측창 90 개장일의 상태·거래대금."""
    ws = wb.create_sheet("2b_L2_원자료")
    r = _title(ws, 1, f"L2 집계 원자료 — {SAMPLE_SID} · 관측 종료일 {REBAL} 기준 최근 90 개장일",
               "코드가 뽑은 상태코드를 그대로 옮긴 것이 아니라, 입력 원자료 + 정지등록만 옮겼다.")
    r = _header(ws, r, ["일자", "원종가", "거래량", "KRX 거래대금", "정지등록", "① 상태코드", "② 반영 거래대금"],
                [12, 12, 14, 18, 10, 16, 18])
    for x in states_code.itertuples():
        r = _row(ws, r, [x.market_date, x.raw_close, x.volume, x.exchange_trading_value,
                         x.halt_mark, None, None],
                 [FILL_GIVEN] * 5 + [FILL_INPUT] * 2)


def sheet_l3(wb, basket):
    """L3 — 구성종목·6셀 가중. ADTV90·시즈닝은 L2 에서 직접 산출한 값을 쓴다(블라인드)."""
    ws = wb.create_sheet("3_L3_구성가중")
    r = _title(ws, 1, f"L3 · 구성종목 판정과 6셀 가중 — {REBAL} 회차",
               "룰북 §9(게이트) · §10(6셀 각 1/6, 셀 내 동일가중) · D-13 ① 대안 A 전부 편입")
    r = _header(ws, r, ["종목", "시장", "테마", "① 내 ADTV90", "② 내 시즈닝일수",
                        "③ 하한 통과?", "④ 시즈닝 통과?", "⑤ 편입?", "⑥ 제외사유", "⑦ 셀", "⑧ 최종비중"],
                [10, 8, 16, 18, 14, 14, 14, 10, 24, 20, 14])
    for b in basket.sort_values(["market", "primary_theme", "security_id"]).itertuples():
        note = FILL_NOTE if b.security_id in ("KTOS", "SPCX", SAMPLE_SID) else FILL_GIVEN
        r = _row(ws, r, [b.security_id, b.market, b.primary_theme] + [None] * 8,
                 [note] * 3 + [FILL_INPUT] * 8)
    r += 1
    ws.cell(r, 1, "시장별 P10 하한 (직접 산출)").font = Font(bold=True, size=10)
    r += 1
    for k in ("KR", "US"):
        r = _row(ws, r, [k, None], [None, FILL_INPUT])
    r += 1
    for t in ["※ 붉은 칸 3종목이 경계사례다. 010120 은 정지·분할, SPCX 는 시즈닝 미달, KTOS 는 하한 근접.",
              "※ ①②는 L2 방식으로 18종목 전부 계산해야 한다. 표본 1종목만 하고 나머지는 넘기지 말 것 —",
              "   P10 은 시장별 전체 분포에서 나오므로 한 종목만으로는 하한 자체를 못 구한다.",
              "※ ⑧ 총합이 정확히 1.0 인지 반드시 검산한다(룰북 §14 비중합계 100%).",
              "※ 빈 셀이 생기면 같은 테마의 타지역 셀로 재배분한다(룰북 §10, D-10 ③)."]:
        ws.cell(r, 1, t).font = Font(size=9, color="52514E")
        r += 1


def sheet_l4(wb, prices, basket, fx, idx_code):
    """L4 — 지수값. KR 다리와 US 다리를 분리해 미결(B-2·B-3) 영향을 격리한다.

    구성종목·비중은 L3 에서 직접 산출한 값을 쓴다. 코드의 편입 결과를 여기 실으면
    L3 의 답(어느 3종목이 빠졌는가)이 그대로 노출되므로 **유니버스 18종목 전부**를 싣는다.
    """
    ws = wb.create_sheet("4_L4_지수")
    r = _title(ws, 1, "L4 · 지수값 검산 (KR 다리 / US 다리 분리)",
               "룰북 §13 1~4항 / 구간 내 리밸 0회 → §13.5 연결계수·제수(B-1) 미발동 = 지금 검산 가능")

    ws.cell(r, 1, "구성종목과 목표비중 — L3 ⑤⑧ 에서 직접 산출한 값을 옮겨 적는다").font = \
        Font(bold=True, size=10)
    r = _header(ws, r + 1, ["종목", "시장", "① 편입?", "② 목표비중"], [10, 8, 12, 16])
    for b in basket.sort_values(["market", "security_id"]).itertuples():
        r = _row(ws, r, [b.security_id, b.market, None, None],
                 [FILL_GIVEN, FILL_GIVEN, FILL_INPUT, FILL_INPUT])

    r += 1
    ws.cell(r, 1, "일별 원자료 — 유니버스 18종목 원종가 (KR 원화 / US 달러) + 환율").font = \
        Font(bold=True, size=10)
    r += 1
    sids = list(basket.sort_values(["market", "security_id"]).security_id)
    r = _header(ws, r, ["일자", "환율(원/달러)"] + sids, [12, 14] + [11] * len(sids))
    for d in INDEX_DAYS:
        vals = [d, float(fx.get(d)) if d in fx.index else None]
        for sid in sids:
            g = prices[(prices.security_id == sid) & (prices.market_date == d)]
            vals.append(float(g.raw_close.iloc[0]) if len(g) else None)
        r = _row(ws, r, vals, [FILL_GIVEN] * (2 + len(sids)))

    r += 1
    ws.cell(r, 1, "수기 산출").font = Font(bold=True, size=10)
    r = _header(ws, r + 1, ["일자", "① KR 다리 수익률", "② US 다리 수익률(원화)",
                            "③ 합산 지수수익률", "④ 지수레벨"],
                [12, 22, 24, 22, 18])
    for d in INDEX_DAYS:
        r = _row(ws, r, [d, None, None, None, None], [FILL_GIVEN] + [FILL_INPUT] * 4)

    r += 1
    for t in ["※ KR 다리부터 하라. 환율이 안 들어가므로 미결(B-2 적용시점·B-3 결측일)에 걸리지 않는다.",
              "※ US 다리는 '당일 환율 적용'을 가정하고 계산한 뒤, 불일치가 나면 그것이 B-2·B-3 때문인지 판별한다.",
              "※ 구간 내 리밸런싱이 0회이므로 연결계수·제수(B-1)는 발동하지 않는다. 효력발생일 대비 가격비만 쓴다.",
              "※ 2026-04-13 은 010120 분할일이다. 규칙대로 원종가를 쓰면 -77% 가 나오고 코드와 '일치'한다.",
              "   → 5_대조표 의 (d) 칸에 '규칙 적용 결과 일치 / 경제적 타당성 불가'로 나눠 적을 것."]:
        ws.cell(r, 1, t).font = Font(size=9, color="D03B3B" if t.startswith("※ 2026") else "52514E")
        r += 1


def build_sealed(ledger, th, idx, weights, path):
    """대조용 정답지 — 코드 산출값만 모은 별도 파일. 수기 채움 완료 후에 연다."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("봉인_코드값")
    r = _title(ws, 1, "J-5 대조용 코드 산출값 (봉인)",
               "★ 수기 워크시트를 모두 채우고 저장한 뒤에 열 것. 먼저 보면 블라인드 검산이 성립하지 않는다.")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24

    led = ledger[(ledger.security_id == SAMPLE_SID)
                 & (ledger.review_cycle_id == f"RC-{REBAL}")]
    code = led.iloc[0] if len(led) else None
    ws.cell(r, 1, f"L2 · {SAMPLE_SID} ADTV90·시즈닝").font = F_T
    r = _header(ws, r + 1, ["항목", "코드 산출값"], [36, 24])
    for label, key in L2_ITEMS:
        v = ""
        if key is not None and code is not None and pd.notna(code[key]):
            v = float(code[key]) if isinstance(code[key], (int, float, np.floating)) else code[key]
        r = _row(ws, r, [label, v], [None, FILL_CODE])

    r += 1
    ws.cell(r, 1, "L3 · 시장별 P10 하한 · 편입 결과").font = F_T
    r = _header(ws, r + 1, ["항목", "코드 산출값"], [36, 24])
    for k, v in th.items():
        r = _row(ws, r, [f"P10 하한 ({k})", v], [None, FILL_CODE])
    r = _row(ws, r, ["편입 종목수", len(weights)], [None, FILL_CODE])
    r = _row(ws, r, ["비중 총합", float(weights.final_target_weight.sum())], [None, FILL_CODE])
    for x in weights.itertuples():
        r = _row(ws, r, [f"  비중 {x.security_id}", float(x.final_target_weight)],
                 [None, FILL_CODE])

    r += 1
    ws.cell(r, 1, "L4 · 일별 지수레벨").font = F_T
    r = _header(ws, r + 1, ["일자", "코드 지수레벨"], [36, 24])
    ic = idx.set_index("market_date")["index_level"]
    for d in INDEX_DAYS:
        r = _row(ws, r, [d, float(ic[d]) if d in ic.index else ""], [None, FILL_CODE])
    wb.save(path)


def sheet_compare(wb):
    ws = wb.create_sheet("5_대조표")
    r = _title(ws, 1, "대조표 — 수기 vs 코드",
               "허용오차는 0_표지 참조. 판정류는 완전일치, 수치는 상대오차 기준.")
    r = _header(ws, r, ["층", "항목", "수기값", "코드값", "차이", "허용오차 내?", "불일치 유형", "비고"],
                [8, 34, 20, 20, 16, 14, 16, 40])
    for layer, items in [("L1", ["04-08 상태코드", "04-08 일별 거래대금", "04-13 상태코드"]),
                         ("L2", ["ADTV90 분모", "공식 ADTV90", "시즈닝 유효관측일수", "KR P10 하한"]),
                         ("L3", ["편입 종목수", "비중 총합", "KTOS 편입 여부", "SPCX 편입 여부"]),
                         ("L4", ["04-13 KR 다리 수익률", "04-13 지수레벨", "04-14 지수레벨"])]:
        for it in items:
            r = _row(ws, r, [layer, it, None, None, None, None, None, None],
                     [None, None] + [FILL_INPUT] * 6)
    r += 2
    ws.cell(r, 1, "불일치 4분류").font = F_T
    r = _header(ws, r + 1, ["유형", "뜻", "처리"], [10, 46, 50])
    for a, b, c in [("(a)", "수기 실수 — 내가 틀림", "재계산, 기록만"),
                    ("(b)", "코드 버그 — 규칙은 명확한데 코드가 다름", "엔진 담당 이관"),
                    ("(c)", "규칙 모호·부재 — 조항이 없거나 두 갈래", "결정로그 미결 등록"),
                    ("(d)", "규칙대로인데 결과가 틀림 — 계산은 조항대로인데 경제적으로 말이 안 됨",
                     "규칙 자체의 결함으로 상신")]:
        r = _row(ws, r, [a, b, c], [FILL_NOTE if a == "(d)" else None, None, None])
    r += 1
    ws.cell(r, 1, "※ (d)를 별도로 두는 이유: 규칙대로 계산하면 수기·코드가 나란히 틀린 값에서 "
                  "'일치'로 찍힌다. 검산이 오류에 도장을 찍어주는 걸 막는 칸이다.").font = \
        Font(size=9, color="D03B3B")
    r += 2
    ws.cell(r, 1, "최종 판정").font = F_T
    r = _header(ws, r + 1, ["항목", "값", "비고"], [24, 20, 60])
    for k in ["validation_status (PASS/WARN/HOLD/FAIL)", "approval_status (팀 검토 후)",
              "검산 수행일", "검산자"]:
        r = _row(ws, r, [k, None, None], [None, FILL_INPUT, FILL_INPUT])


# ─────────────────────────────────────────────────────────────
def main():
    inp, out = P.PILOT_INPUT, P.PILOT_OUTPUT
    P.require(os.path.join(inp, "prices.csv"), "엔진 입력")
    os.makedirs(HERE, exist_ok=True)

    snap = snapshot(inp, os.path.join(HERE, "snapshot_hashes.csv"))
    try:
        import subprocess
        commit = subprocess.run(["git", "rev-parse", "HEAD"], capture_output=True,
                                text=True, cwd=P.ROOT).stdout.strip()[:12]
    except Exception:
        commit = ""

    rd = lambda d, n: pd.read_csv(os.path.join(d, n), dtype={"security_id": str})
    prices, basket = rd(inp, "prices.csv"), rd(inp, "seed_basket.csv")
    halts, cal = rd(inp, "halts.csv"), pd.read_csv(os.path.join(inp, "calendar.csv"))
    fx = pd.read_csv(os.path.join(inp, "fx.csv")).set_index("market_date")["fx_rate"]
    ledger, weights = rd(out, "adtv90_ledger.csv"), rd(out, f"weights_{REBAL}.csv")
    idx = pd.read_csv(os.path.join(out, "index_vs_benchmark.csv"))
    th = json.load(open(os.path.join(out, f"thresholds_{REBAL}.json"), encoding="utf-8"))["provisional_P10"]

    # L2 원자료: 관측 종료일 기준 최근 90 개장일 (상태코드는 사람이 판정하도록 비운다)
    kr_open = sorted(cal[(cal.market == "KR") & (cal.is_market_open == 1)]
                     ["market_date"].astype(str))
    hs = {str(x.market_date) for x in halts.itertuples()
          if x.security_id == SAMPLE_SID and str(x.full_day_halt).strip().lower() in ("true", "1", "y")}
    win = [d for d in kr_open if d <= REBAL][-90:]
    p1 = prices[prices.security_id == SAMPLE_SID].set_index("market_date")
    l2raw = pd.DataFrame([{
        "market_date": d,
        "raw_close": float(p1.at[d, "raw_close"]) if d in p1.index else None,
        "volume": float(p1.at[d, "volume"]) if d in p1.index else None,
        "exchange_trading_value": float(p1.at[d, "exchange_trading_value"])
        if d in p1.index and pd.notna(p1.at[d, "exchange_trading_value"]) else None,
        "halt_mark": "O" if d in hs else ""} for d in win])

    wb = Workbook()
    wb.remove(wb.active)
    sheet_cover(wb, snap, commit)
    sheet_l1(wb, prices, halts, cal)
    sheet_l2(wb)
    sheet_l2_raw(wb, l2raw)
    sheet_l3(wb, basket)
    sheet_l4(wb, prices, basket, fx, idx)
    sheet_compare(wb)

    path = os.path.join(HERE, "J5_수기검산_워크시트.xlsx")
    sealed = os.path.join(HERE, "J5_코드값_봉인.xlsx")
    wb.save(path)
    build_sealed(ledger, th, idx, weights, sealed)
    print(f"[J-5] 스냅샷 {len(snap)}파일 해시 기록 → snapshot_hashes.csv")
    print(f"[J-5] 워크시트 7시트 생성 → {os.path.relpath(path, P.ROOT)}")
    print(f"      주 표본 {SAMPLE_SID} × {SAMPLE_DAYS[0]}~{SAMPLE_DAYS[-1]} · 리밸 {REBAL} · "
          f"L2 원자료 {len(l2raw)}일")
    print(f"[J-5] 대조용 코드값 봉인 → {os.path.relpath(sealed, P.ROOT)}")
    print("      워크시트에 코드 산출값은 없다(블라인드). 수기 칸을 모두 채우고 저장한 뒤에")
    print("      봉인 파일을 열어 5_대조표에 옮겨 적을 것. 엑셀 수식만 사용(파이썬 금지).")


if __name__ == "__main__":
    main()
